# -*- coding: utf-8 -*-
"""
整合型高速量化分析引擎 (In-Memory 資料處理 + V18 HorseFinder Pro 版)
====================================================================
V17 架構 100% 保留：
1. 負責讀取 Google Sheet 股票清單。
2. 於記憶體內即時下載 Yahoo 價量與 FinMind 籌碼資料 (統一 UTC-aware 時間軸)。
3. 計算完整核心指標 (波動壓縮、趨勢加速度、RSI背離等高級因子)。
4. 全球風險因子：NASDAQ / SOX / VIX / DXY / US10Y。
5. 量化 Macro Overlay：連續加權風險分數(0~100) + Z-score 正規化。
6. 市場 Regime 分流：NORMAL / CAUTION / HIGH_RISK / CRASH 四態切換。
7. Backtest 含動態 Position Sizing：風險越高倉位越小。
8. Feature 過多自動警告 + 強制上限控制。
9. 波段交易專用 Excel 排版：波段總體看板 + 動態 10EMA 移動停利指標。

V18 HorseFinder Pro 新增（純新增，不更動上述 V17 流程）：
10. 【BB 模組】BB Width / BB Width Change / BB Width ZScore / BB Squeeze / 強化 BB %B / BB Confidence Bonus。
11. 【HorseFinder】獨立黑馬評分機制 (Horse Score 0~100，>=70 列為黑馬候選股，不直接影響 AI 分數)。
12. 【Black Horse Rank】新增 Excel 分頁，只列最有爆發力的前 20 檔黑馬候選股。
13. 【AI 信心等級】★ 星等信心分數，個股頁與總表皆顯示 AI Score / Confidence / Horse / Regime / Risk / BB Squeeze / Macro。
14. 【SHAP 特徵重要度】全域 SHAP 重要性納入因子排名，並提供個股當次預測的局部 SHAP 解釋。
15. 【Walk-Forward Validation】既有 anchored expanding-window 回測正式標註並擴充折數與報酬標準差統計。
16. 【Optuna 自動調參】以貝葉斯優化取代/加強 RandomizedSearchCV，未安裝 Optuna 時自動退回 V17 版本。
17. 【Kelly Position Sizing】依個股歷史真實交易紀錄估算半凱利倉位，與 ATR / 風險倉位取最小值聯合決策。

V18.1 決策層升級（依專家建議追加）：
18. 【Horse Score 星等化】黑馬候選欄位由單純 YES/NO 改為 5 級星等（90+★★★★★／70+★★★★☆／50+★★★☆☆／30+★★☆☆☆／<30★☆☆☆☆），資訊更連續。
19. 【AI Decision Gate】AI上漲機率 <50% 時，決策最高鎖定「觀察」，除非 HorseScore >= 90 才允許 Override 為「搶先布局」，避免「AI沒信心卻建議買入」的過度積極矛盾。
20. 【Kelly Fallback】樣本不足時不再顯示生硬的 N/A，改依 AI 上漲機率分級（>=70:20%／60~70:15%／50~60:10%／<50:5%）給出保守 Fallback 建議，並明確標註為 Fallback。
21. 【Final Decision Engine】把 AI Decision Gate、HorseFinder、風險模式(Regime) 統整成單一「最終決策」欄位/面板，避免報表中不同欄位互相矛盾。

V18.2 依專家建議再升級（本次修改）：
22. 【動態 AI Gate 門檻】AI Decision Gate 門檻不再固定 50%，改依「類股別 + 趨勢強度 + 歷史Sharpe」動態調整
    （科技股 55% / ETF 50% / 傳產 50% / 金融股 45%，高趨勢+高Sharpe再降5%，下限43%），避免錯過低波動金融股的主升段。
23. 【階梯式 Horse Override】HorseScore 覆寫 AI Gate 不再單一鎖 90 分，改為三段式（90分:AI機率>=30%可覆寫／
    80分:AI機率>=40%可覆寫／70分:AI機率>=45%可覆寫），並將星等意義對齊「70准黑馬／80強黑馬／90爆發黑馬」。
24. 【資金流因子】新增「成交值 / 20日均成交值」爆量倍率（>2.5倍視為法人進場訊號），納入 HorseScore 評分，
    比單純技術指標更早反映主力進場動作；法人（外資/投信）連買天數仍保留於評分中。
25. 【分業風險模型】全球風險模式改依類股別切換風險來源：AI電子股看 SOX/NASDAQ／金融股看美債殖利率
    +台幣匯率+金融類股指數／傳產股看油價+中國(FXI代理)+美元指數，取代單一套用 NASDAQ/SOX 對所有類股。
26. 【市場 Regime 分類 MARKET_MODE】不再所有股票用同一套邏輯：把既有 macro_regime(7級) + VIX
    收斂成 BULL/SIDEWAYS/BEAR/PANIC 四種市場模式，各模式有各自「偏好類股」與「偏好策略風格」
    （多頭→突破動能／盤整→均值回歸RSI低接／空頭→防守金融逆勢／恐慌→現金為王），
    並讓 AI Decision Gate 門檻依模式再微調（多頭放寬、空頭收緊、恐慌等同鎖死），
    同時對「風格與模式不match」的訊號（例如空頭市場硬做非金融股突破）做降級提示，避免勝率暴跌、回撤增加。

V18.3 依專家建議再升級（本次修改）：
27. 【資金輪動 Sector Heat Score】新增細分類股（AI/半導體/記憶體/PCB/散熱/網通/航運/金融/傳產）+
    Sector Heat Score（0~100，以本次監控清單中同類股近5日「價格動能＋量能擴張」的相對強弱計算，
    例如 AI=92／PCB=88／金融=42），解決「AI→PCB→散熱→金融→航運」資金輪動時、只看單股容易
    「剛買進、資金已撤退」的問題。個股 FinalScore = 個股綜合分數 × 類股熱度倍率(0.6~1.4)，
    熱度<35分（資金退潮警示）時，對積極買進訊號額外標註輪動風險提示；新增「Sector Heat」分頁
    列出各類股熱度排行，並將總表排序依據由 composite_score 改為 FinalScore(含熱度加權)。

V18.4 依專家建議再升級（本次修改）：
28. 【Profit Factor / Expectancy】專家建議：勝率不是最重要的指標，Profit Factor(獲利因子) 與
    Expectancy(單筆期望值) 才是專業交易更重要的指標。新增 compute_trade_stats()：
      Gross Profit / Gross Loss → Profit Factor = Gross Profit ÷ Gross Loss（無虧損交易時顯示∞）
      Expectancy(E) = 勝率×平均獲利 - 敗率×平均虧損 = 全部交易報酬平均值（例如 E=+2.8% 代表平均每筆交易賺2.8%）
    個股層級：寫入每檔個股工作表的「波段回測績效」面板；總表層級：新增「💎Profit Factor」「💎Expectancy」
    兩欄，並在總表看板新增「Profit Factor」看板卡片、把原本「平均波段單筆報酬」正式更名並標註為
    Expectancy；「真實回測總表」分頁底部新增整體 Profit Factor / Expectancy / Gross Profit-Gross Loss 彙總列。
"""

import os
import re
import time
import warnings
import io
import requests
from datetime import datetime, timedelta
import pandas as pd
import numpy as np

# 嘗試匯入必要量化套件
try:
    import yfinance as yf
    from FinMind.data import DataLoader
    
    from ta.momentum import RSIIndicator, StochasticOscillator
    from ta.trend import MACD, EMAIndicator, SMAIndicator, ADXIndicator
    from ta.volatility import BollingerBands, AverageTrueRange
    from ta.volume import OnBalanceVolumeIndicator

    import lightgbm as lgb
    import xgboost as xgb
    from sklearn.model_selection import TimeSeriesSplit, RandomizedSearchCV
    from sklearn.metrics import mean_squared_error
    from sklearn.feature_selection import mutual_info_regression
    from scipy.stats import pearsonr, spearmanr

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.chart import LineChart, Reference

except ImportError as e:
    print(f"請先安裝必要套件: pip install pandas numpy yfinance FinMind ta lightgbm xgboost scikit-learn openpyxl requests")
    print(f"匯入失敗詳細資訊: {e}")
    import sys
    sys.exit(1)

# ========= V18：選用套件（缺少時自動降級，不影響主流程） =========
try:
    import shap
    HAS_SHAP = True
except ImportError:
    shap = None
    HAS_SHAP = False
    print("  ℹ️ 未安裝 shap，SHAP 特徵解釋將自動略過（不影響其餘功能）。可執行 pip install shap 啟用。")

try:
    import optuna
    optuna.logging.set_verbosity(optuna.logging.WARNING)
    HAS_OPTUNA = True
except ImportError:
    optuna = None
    HAS_OPTUNA = False
    print("  ℹ️ 未安裝 optuna，將自動退回 V17 RandomizedSearchCV 調參。可執行 pip install optuna 啟用貝葉斯調參。")

warnings.filterwarnings("ignore")

# ================= 設定區 =================
# Google Sheet 串接設定
SHEET_ID       = "1iyS553YvCHKwLdOAeFzghJASLAY_NVitPjRky0CzQLI"
GID            = "630045424"
SHEET_TAB_NAME = "權值股"

# 資料下載範圍 (全記憶體處理，拉取2年長天期資料)
PERIOD         = "2y"
FINMIND_TOKEN  = ""  # 若您有 FinMind Token 請填入此處

# V14 核心設定
VERSION        = "V18.4_HorseFinder_ProfitFactor"
FUTURE_DAYS    = 5   
HORIZONS       = [1, 3, 5, 20] 
TOP_N_FEATURES = 10  
MIN_IMPORTANCE = 0.01

# ========= V18 HorseFinder 黑馬評分權重（獨立於 AI 模型分數之外） =========
# V18.2：加入「成交值爆量倍率」資金流因子，並將 volume_breakout(股數爆量) 權重下修，
#        避免與 value_surge_ratio(成交值/金額爆量) 重複計分過重；總分仍為 100。
HORSE_SCORE_WEIGHTS = {
    "bb_squeeze":        20,   # BB Squeeze（波動壓縮到位）
    "bb_width_low10pct": 15,   # BB Width < 歷史10%分位
    "volume_breakout":   15,   # 成交量(股數)突破 2 倍均量
    "value_surge_ratio": 15,   # 【V18.2新增】成交值 / 20日均成交值 > 2.5倍，代表法人資金進場，常比技術指標更早反應
    "new_high_20d":      15,   # 20 日新高
    "macd_golden_cross": 10,   # MACD 黃金交叉（近3日內）
    "adx_strong":        10,   # ADX > 25
    "institution_streak":10,   # 法人（外資或投信）連買 >= 3 日
}
VALUE_SURGE_RATIO_THRESHOLD = 2.5   # 【V18.2新增】成交值/20日均成交值 倍率門檻，超過視為法人資金進場訊號
HORSE_SCORE_THRESHOLD = 70     # >= 此分數列為「准黑馬」候選股（進入 Black Horse Rank 分頁 / AI Decision Gate 判斷用）

# V18.2：Horse Score 星等 / 分級意義對齊專家建議 —— 70准黑馬／80強黑馬／90爆發黑馬，取代原本「90才有意義」的單一門檻
HORSE_STAR_TIERS = [
    (90, "★★★★★", "爆發黑馬"),
    (80, "★★★★☆", "強黑馬"),
    (70, "★★★☆☆", "准黑馬"),
    (50, "★★☆☆☆", "中等"),
    (0,  "★☆☆☆☆", "普通"),
]

# ========= V18.2：AI Decision Gate 動態門檻設定（取代固定 50%） =========
# 不同類股波動特性不同：金融股波動低，48~55%的AI機率其實已經足夠操作，固定50%反而常錯過金融股主升段。
AI_GATE_LOW_PROB   = 50   # 通用預設門檻（找不到類股分類、或當作 fallback 時使用）
AI_GATE_THRESHOLD_BY_CATEGORY = {
    "科技":   55,   # 高波動科技/電子股：AI 需要更高信心才放行，避免追高假突破
    "電子":   55,
    "ETF":    50,   # ETF：波動介於中間，維持通用門檻
    "傳產":   50,
    "金融":   45,   # 低波動金融股：確信度不需要那麼高即可視為足夠
    "default":50,
}
# 若「趨勢分數」與「歷史Sharpe」同時偏高（強趨勢股），門檻可再放寬，讓模型更像真人交易員判斷「趨勢對了，AI只是保守」
AI_GATE_TREND_SHARPE_DISCOUNT = {
    "trend_score_min": 70,   # 趨勢分數（0~100，由 ADX與多週期方向一致性換算）達此值視為「超強趨勢」
    "sharpe_min":       1.0, # 歷史 Walk-Forward Sharpe 達此值視為「高品質趨勢」
    "discount":         5,   # 兩者同時滿足時，門檻再降 5 個百分點（例：金融股 45% → 40%，但仍受下限保護）
}
AI_GATE_THRESHOLD_FLOOR = 43   # 動態門檻無論如何調整，最低不得低於此值（超強趨勢股 43~45% 區間）

# ========= V18.2b：市場 Regime 分類（MARKET_MODE） =========
# 專家建議：現在模型「所有股票用同一套邏輯」，但多頭/空頭/盤整適合的策略完全不同
# （多頭→突破動能／科技股；空頭→防守/金融/高殖利率/逆勢；盤整→均值回歸/RSI低接/區間策略）。
# 若在空頭市場仍硬做突破股，勝率會暴跌、回撤增加。
# 這裡把既有的 process_macro_data() 產出的 7 級 macro_regime（強趨勢多頭～崩盤）+ VIX，
# 收斂成 4 種 MARKET_MODE：BULL / SIDEWAYS / BEAR / PANIC，並讓 AI Gate 門檻與訊號依模式切換。
MARKET_MODE_MAP = {
    "強趨勢多頭": "BULL", "緩漲盤": "BULL",
    "高波動震盪": "SIDEWAYS", "低波動盤整": "SIDEWAYS",
    "空頭反彈": "BEAR", "主跌段": "BEAR",
    "崩盤": "PANIC",
}
MARKET_MODE_PLAYBOOK = {
    "BULL": {
        "label": "🐂多頭",
        "favored_categories": ["科技", "電子", "ETF"],
        "favored_style": "breakout / momentum（順勢突破、動能股）",
        "ai_gate_delta": -3,     # 多頭順風，AI Gate 門檻可再放寬3個百分點
    },
    "SIDEWAYS": {
        "label": "🦀盤整",
        "favored_categories": ["金融", "傳產", "ETF"],
        "favored_style": "mean-reversion / RSI低接（區間高出低進，不追突破）",
        "ai_gate_delta": +3,     # 盤整盤追突破容易假突破，門檻略提高
    },
    "BEAR": {
        "label": "🐻空頭",
        "favored_categories": ["金融"],
        "favored_style": "defensive / 高殖利率 / 逆勢低接（防守為主）",
        "ai_gate_delta": +5,     # 空頭市場，訊號需要更高AI信心才放行
    },
    "PANIC": {
        "label": "🚨恐慌",
        "favored_categories": [],
        "favored_style": "cash-is-king（現金為王，全面防禦）",
        "ai_gate_delta": +100,   # 實務上等同鎖死，AI Gate 幾乎不可能被滿足
    },
}
MARKET_MODE_RSI_OVERSOLD = 30   # BEAR/SIDEWAYS 模式下，RSI低於此值可視為「逆勢低接」候選，不強制降級

# V18.2：階梯式 HorseScore Override —— 取代「只有 >=90 才覆寫」的單一硬門檻，改為分數越高、可容忍的AI機率下限越低
# 由高到低比對，只要 (horse_score, ai_prob) 同時達標即允許 Override 為「搶先布局」
AI_GATE_OVERRIDE_TIERS = [
    (90, 30),   # 爆發黑馬：AI機率 >= 30% 即可覆寫（訊號極強，AI保守也值得卡位）
    (80, 40),   # 強黑馬：AI機率 >= 40% 可覆寫
    (70, 45),   # 准黑馬：AI機率 >= 45% 可覆寫
]
AI_GATE_OVERRIDE_HORSE = AI_GATE_OVERRIDE_TIERS[0][0]  # 向下相容：保留舊變數名稱，供其他地方參照最高階門檻

KELLY_FALLBACK_TIERS = [   # V18.1：Kelly 樣本不足時的備用倉位（依 AI 上漲機率分級）
    (70, 0.20), (60, 0.15), (50, 0.10), (0, 0.05),
]
HORSE_RANK_TOP_N       = 20    # Black Horse Rank 分頁列出前 N 名

# ========= V18.2：股票類股分類（供動態 AI Gate 門檻 / 分業風險模型使用） =========
# Google Sheet 股票清單只有代號與公司名稱，沒有類股欄位，因此改用「代號區間 + 公司名稱關鍵字」做輕量分類。
# 此分類為輔助判斷用途，非精確產業分類；若之後 Google Sheet 有正式類股欄位，可直接改讀該欄位取代本函式。
ETF_CODE_PREFIXES = ("0050", "0051", "0052", "0053", "0055", "0056", "0057", "0061",
                     "006", "00")   # 006xx / 00xx 開頭代碼多為 ETF（如 0050、0056、00878、006208...）
FINANCE_KEYWORDS = ["金控", "銀行", "保險", "證券", "票券", "產險", "人壽", "金融控股", "投信"]
TECH_KEYWORDS = ["半導體", "電子", "科技", "晶", "光電", "通訊", "網通", "資訊", "IC", "面板",
                  "軟體", "電腦", "手機", "光學", "封測", "晶圓", "PCB", "電路板"]
TRADITIONAL_KEYWORDS = ["塑膠", "鋼鐵", "水泥", "紡織", "食品", "營建", "航運", "石化", "橡膠",
                         "造紙", "玻璃", "汽車", "化學", "農業", "食"]

# ========= V18.3：資金輪動 Sector Heat Score（細分類股 + 熱度評分） =========
# 專家建議：台股常見「AI → PCB → 散熱 → 金融 → 航運」輪動，若只看單股，剛買進時資金可能已撤退到下一個類股。
# 這裡用比 classify_stock_category 更細的「主題型」關鍵字表做分類，供資金輪動熱度評分使用（比對順序：由細到粗）。
SECTOR_FINE_KEYWORDS = {
    "AI":     ["人工智慧", "AI", "伺服器", "繪圖處理", "資料中心"],
    "半導體": ["半導體", "晶圓", "IC設計", "晶片", "封測"],
    "記憶體": ["記憶體", "DRAM", "快閃", "儲存"],
    "PCB":    ["PCB", "電路板", "覆銅", "載板"],
    "散熱":   ["散熱", "導熱", "均熱片", "液冷", "水冷"],
    "網通":   ["網通", "通訊", "路由器", "基地台", "光纖"],
    "航運":   ["航運", "海運", "貨櫃", "航空", "船務"],
    "金融":   FINANCE_KEYWORDS,
    "傳產":   TRADITIONAL_KEYWORDS,
}
SECTOR_HEAT_LOOKBACK_DAYS   = 5     # 資金輪動熱度：以近5日「價格動能＋量能擴張」衡量，比技術指標更早反映資金流向
SECTOR_HEAT_MIN_MEMBERS     = 2     # 該細分類股本次至少要有N檔股票同時在監控清單，才單獨計算熱度分數，否則併入 default(50分中性)
SECTOR_HEAT_MULTIPLIER_MIN  = 0.6   # FinalScore = 個股分數 × 類股熱度倍率，倍率下限（避免冷門股分數被打到0）
SECTOR_HEAT_MULTIPLIER_MAX  = 1.4   # 倍率上限（避免熱門股分數被過度放大）
SECTOR_HEAT_COLD_THRESHOLD  = 35    # 熱度分數低於此值視為「資金退潮」警示區，對積極訊號額外註記風險
SECTOR_HEAT_HOT_THRESHOLD   = 70    # 熱度分數高於此值視為「資金流入熱區」

# ========= V18 Kelly Position Sizing 參數 =========
KELLY_MIN_RECORDS = 5      # 至少需要多少筆真實回測紀錄才計算 Kelly
KELLY_HALF_FACTOR = 0.5    # 半凱利係數（保守化，避免過度槓桿）
KELLY_MAX_FRACTION = 0.25  # Kelly 倉位硬上限

TARGET_CLIP_MIN = -15.0
TARGET_CLIP_MAX = 15.0
CLS_THRESHOLD   = 0.0  

COST_PCT = 0.3  # 實戰交易摩擦成本 (含滑價)

# ========= V15 全球風險設定 =========
US_RISK_TICKERS = {
    "nasdaq": "^IXIC",    # NASDAQ 綜合指數
    "sox":    "SOXX",     # 費城半導體 ETF (SOX)
    "vix":    "^VIX",     # CBOE 恐慌指數
    "dxy":    "DX-Y.NYB", # 美元指數
    "us10y":  "^TNX",     # 美國10年期公債殖利率
    # ── V18.2 新增：分業風險模型所需的額外因子 ──────────────────
    "oil":    "CL=F",     # WTI 原油期貨（傳產類股成本/景氣代理）
    "twd":    "TWD=X",    # 美元/新台幣匯率（金融股與出口傳產敏感度高）
    "china":  "FXI",      # iShares中國大型股ETF（作為中國景氣/PMI的市場代理指標）
    "tw_fin": "0055.TW",  # 元大MSCI金融ETF（台股金融類股代理指數）
}
VIX_PANIC_THRESHOLD   = 25.0   # VIX 超過此值視為恐慌市場
VIX_FEAR_THRESHOLD    = 20.0   # VIX 超過此值進入警戒
NASDAQ_DROP_WARN      = -2.0   # NASDAQ 單日跌幅觸發警示 (%)
NASDAQ_DROP_SEVERE    = -3.5   # NASDAQ 單日跌幅觸發嚴重警示 (%)
SOX_DROP_WARN         = -3.0   # SOX 單日跌幅觸發警示 (%)
SOX_DROP_SEVERE       = -5.0   # SOX 嚴重下跌門檻 (%)
US10Y_JUMP_WARN       = 0.10   # 美債殖利率單日跳升 (%)
USD_TWD_RISK          = 32.5   # 美元/台幣匯率風險水位

# ========= V16 量化風控參數 =========
MAX_FEATURES          = 25     # 特徵數量硬上限（防止 overfitting）
POSITION_SIZE_BASE    = 0.20   # 基礎最大倉位（正常市場）
POSITION_SIZE_CAUTION = 0.12   # 警戒市場倉位上限
POSITION_SIZE_HIGHRISK= 0.06   # 高風險市場倉位上限
POSITION_SIZE_CRASH   = 0.00   # 崩盤市場倉位（全現金）
RISK_SCORE_WEIGHTS = {         # 量化 Risk Score 各因子權重（加總=1.0，通用/預設模型，科技電子股沿用此表）
    "vix":      0.30,
    "nasdaq":   0.25,
    "sox":      0.25,
    "us10y":    0.10,
    "dxy":      0.10,
}

# ========= V18.2：分業風險模型 —— 不同類股對「風險來源」的敏感度不同，不應全部套用 NASDAQ/SOX =========
# 專家建議：金融股（如2882）與 SOX 相關性低，更該看美債殖利率／匯率／金融指數；
#          傳產股更該看油價／中國景氣；只有 AI電子/科技股才適合沿用 SOX/NASDAQ 為主的風險模型。
SECTOR_RISK_WEIGHTS = {
    "科技": {"vix": 0.30, "nasdaq": 0.30, "sox": 0.30, "us10y": 0.05, "dxy": 0.05},
    "電子": {"vix": 0.30, "nasdaq": 0.30, "sox": 0.30, "us10y": 0.05, "dxy": 0.05},
    "金融": {"vix": 0.15, "us10y": 0.30, "dxy": 0.20, "twd": 0.20, "tw_fin": 0.15},
    "傳產": {"vix": 0.15, "oil": 0.30, "china": 0.30, "dxy": 0.15, "nasdaq": 0.10},
    "ETF":  RISK_SCORE_WEIGHTS,
    "default": RISK_SCORE_WEIGHTS,
}

COLORS = {
    "header_bg":  "1F3864", "header_fg":  "FFFFFF", "subheader":  "2E75B6",
    "green_bg":   "E2EFDA", "alt_row":    "EBF3FB", "border":     "BDD7EE", "summary_bg": "FFF2CC",
}
INVALID_SHEET_CHARS = r'\/*?:[]'
# ==========================================


# ==========================================
# 🌐 V15：全球風險因子下載與計算
# ==========================================

def fetch_us_risk_data() -> dict:
    """
    下載美股風險指標最新資料，回傳含昨日收盤與日報酬的 dict。
    使用 yfinance 抓取過去 5 個交易日資料取最近值。
    """
    risk = {
        "nasdaq_ret1": 0.0, "nasdaq_ret3": 0.0,
        "sox_ret1": 0.0,    "sox_ret5": 0.0,
        "vix_close": 18.0,  "vix_change": 0.0,
        "dxy_ret1": 0.0,    "us10y_close": 4.0,
        "us10y_jump": 0.0,  "fetch_ok": False,
        # ── V18.2 新增：分業風險模型因子（金融/傳產股用） ──
        "oil_ret1": 0.0, "twd_ret1": 0.0, "china_ret1": 0.0, "tw_fin_ret1": 0.0,
    }
    try:
        print("  📡 正在下載全球風險因子 (NASDAQ/SOX/VIX/DXY/US10Y/油價/匯率/金融指數)...")
        for key, ticker in US_RISK_TICKERS.items():
            df = yf.download(ticker, period="10d", auto_adjust=True, progress=False)
            if df.empty:
                continue
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = df.columns.get_level_values(0)
            df.columns = [c.lower() for c in df.columns]
            df = df.sort_index()
            close = df["close"].dropna()
            if len(close) < 2:
                continue

            if key == "nasdaq":
                risk["nasdaq_ret1"] = float((close.iloc[-1] / close.iloc[-2] - 1) * 100)
                risk["nasdaq_ret3"] = float((close.iloc[-1] / close.iloc[-4] - 1) * 100) if len(close) >= 4 else risk["nasdaq_ret1"]
            elif key == "sox":
                risk["sox_ret1"] = float((close.iloc[-1] / close.iloc[-2] - 1) * 100)
                risk["sox_ret5"] = float((close.iloc[-1] / close.iloc[-6] - 1) * 100) if len(close) >= 6 else risk["sox_ret1"]
            elif key == "vix":
                risk["vix_close"]  = float(close.iloc[-1])
                risk["vix_change"] = float(close.iloc[-1] - close.iloc[-2])
            elif key == "dxy":
                risk["dxy_ret1"] = float((close.iloc[-1] / close.iloc[-2] - 1) * 100)
            elif key == "us10y":
                risk["us10y_close"] = float(close.iloc[-1])
                risk["us10y_jump"]  = float(close.iloc[-1] - close.iloc[-2])
            elif key == "oil":
                risk["oil_ret1"] = float((close.iloc[-1] / close.iloc[-2] - 1) * 100)
            elif key == "twd":
                risk["twd_ret1"] = float((close.iloc[-1] / close.iloc[-2] - 1) * 100)
            elif key == "china":
                risk["china_ret1"] = float((close.iloc[-1] / close.iloc[-2] - 1) * 100)
            elif key == "tw_fin":
                risk["tw_fin_ret1"] = float((close.iloc[-1] / close.iloc[-2] - 1) * 100)

        risk["fetch_ok"] = True
        print(f"  ✅ 全球風險因子下載完成 | NASDAQ:{risk['nasdaq_ret1']:+.2f}% SOX:{risk['sox_ret1']:+.2f}% VIX:{risk['vix_close']:.1f} DXY:{risk['dxy_ret1']:+.2f}% US10Y跳升:{risk['us10y_jump']:+.3f}% | "
              f"油價:{risk['oil_ret1']:+.2f}% 台幣:{risk['twd_ret1']:+.2f}% 中國代理:{risk['china_ret1']:+.2f}% 金融指數:{risk['tw_fin_ret1']:+.2f}%")
    except Exception as e:
        print(f"  ⚠️ 全球風險因子下載失敗，使用中性值：{e}")
    return risk


def compute_global_risk_score(risk: dict) -> tuple:
    """
    V16 升級：量化連續加權風險分數 (0-100) + Z-score 正規化。
    
    各因子以「偏離正常值的 Z-score」計算，加權合成 0~100 分：
      0~25  → NORMAL   (正常市場)
      25~50 → CAUTION  (輕度警戒)
      50~75 → HIGH_RISK(高度警戒)
      75~100→ CRASH    (崩盤模式)
    
    參考專家建議：
      risk_score = 0.30*VIX_z + 0.25*NASDAQ_z + 0.25*SOX_z
                 + 0.10*US10Y_z + 0.10*DXY_z
    """
    details = []

    # ── 各因子 Z-score 正規化（基於歷史正常範圍） ───────────────
    # VIX：正常中心=15, 1σ=5；越高越危險
    vix_z   = max(0, (risk["vix_close"] - 15.0) / 5.0)

    # NASDAQ 日報酬：正常中心=0, 1σ=1%；跌越多越危險（取反）
    nasdaq_z = max(0, -(risk["nasdaq_ret1"] - 0.0) / 1.0)

    # SOX：正常中心=0, 1σ=1.5%；跌越多越危險（取反）
    sox_z    = max(0, -(risk["sox_ret1"] - 0.0) / 1.5)

    # US10Y 單日跳升：正常=0, 1σ=0.05%；跳升越多越危險
    us10y_z  = max(0, (risk["us10y_jump"] - 0.0) / 0.05)

    # DXY 日報酬：正常中心=0, 1σ=0.3%；走強越多越危險
    dxy_z    = max(0, (risk["dxy_ret1"] - 0.0) / 0.3)

    # 3日 NASDAQ 趨勢加成（持續下跌更危險）
    nasdaq3_z = max(0, -(risk["nasdaq_ret3"] - 0.0) / 2.0)

    # ── 加權合成連續風險分數 (raw, 可能 > 1) ────────────────────
    w = RISK_SCORE_WEIGHTS
    raw_score = (
        w["vix"]    * min(vix_z,    3.0)  # cap at 3σ
      + w["nasdaq"] * min(nasdaq_z, 3.0)
      + w["sox"]    * min(sox_z,    3.0)
      + w["us10y"]  * min(us10y_z,  3.0)
      + w["dxy"]    * min(dxy_z,    3.0)
    )
    # 3日趨勢加成（最多+0.3）
    raw_score = min(raw_score + nasdaq3_z * 0.1, 3.5)

    # ── 映射到 0~100 分（3σ = 100分） ──────────────────────────
    continuous_score = round(min(raw_score / 3.0 * 100, 100), 1)

    # ── 離散 Regime 判斷 ────────────────────────────────────────
    if continuous_score >= 75:
        regime = "CRASH"
    elif continuous_score >= 50:
        regime = "HIGH_RISK"
    elif continuous_score >= 25:
        regime = "CAUTION"
    else:
        regime = "NORMAL"

    # ── 組裝描述文字 ────────────────────────────────────────────
    if vix_z > 0.5:
        details.append(f"VIX={risk['vix_close']:.1f}(z={vix_z:.1f}σ)")
    if nasdaq_z > 0.5:
        details.append(f"NASDAQ={risk['nasdaq_ret1']:+.1f}%(z={nasdaq_z:.1f}σ)")
    if sox_z > 0.5:
        details.append(f"SOX={risk['sox_ret1']:+.1f}%(z={sox_z:.1f}σ)")
    if us10y_z > 0.5:
        details.append(f"US10Y跳升{risk['us10y_jump']:+.3f}%(z={us10y_z:.1f}σ)")
    if dxy_z > 0.5:
        details.append(f"DXY={risk['dxy_ret1']:+.2f}%(z={dxy_z:.1f}σ)")

    detail_str = "、".join(details) if details else "全球市場正常"
    return continuous_score, regime, detail_str


def compute_sector_risk_score(risk: dict, category: str = "default") -> tuple:
    """
    V18.2：分業風險模型 —— 依股票類股別，切換風險因子與權重，取代單一套用 NASDAQ/SOX 給所有類股。
      科技/電子股 → 沿用 SOX / NASDAQ / VIX（原始邏輯）
      金融股      → 改看美債殖利率(US10Y) / 台幣匯率(TWD) / 金融類股指數(0055.TW) / DXY
      傳產股      → 改看油價(WTI) / 中國景氣代理(FXI) / DXY
      ETF/其他    → 沿用通用模型（等同 compute_global_risk_score）

    回傳 (continuous_score 0~100, regime, detail_str)，regime 判斷區間與 compute_global_risk_score 一致。
    """
    weights = SECTOR_RISK_WEIGHTS.get(category, SECTOR_RISK_WEIGHTS["default"])
    details = []

    # ── 各因子 Z-score 正規化（基於歷史正常範圍，中心值/1σ 皆為經驗估計） ──
    z_components = {}
    if "vix" in weights:
        z_components["vix"] = max(0, (risk.get("vix_close", 15.0) - 15.0) / 5.0)
    if "nasdaq" in weights:
        z_components["nasdaq"] = max(0, -(risk.get("nasdaq_ret1", 0.0)) / 1.0)
    if "sox" in weights:
        z_components["sox"] = max(0, -(risk.get("sox_ret1", 0.0)) / 1.5)
    if "us10y" in weights:
        z_components["us10y"] = max(0, (risk.get("us10y_jump", 0.0)) / 0.05)
    if "dxy" in weights:
        z_components["dxy"] = max(0, (risk.get("dxy_ret1", 0.0)) / 0.3)
    if "twd" in weights:
        # 台幣兌美元單日波動：中心=0, 1σ=0.3%；貶值(twd_ret1>0代表美元計價的TWD變貴=台幣貶)越多風險越高
        z_components["twd"] = max(0, (risk.get("twd_ret1", 0.0)) / 0.3)
    if "tw_fin" in weights:
        # 金融類股指數：中心=0, 1σ=1.0%；下跌越多風險越高
        z_components["tw_fin"] = max(0, -(risk.get("tw_fin_ret1", 0.0)) / 1.0)
    if "oil" in weights:
        # 油價：中心=0, 1σ=2.5%；不論暴漲(成本上升)或暴跌(景氣疑慮)都視為風險，取絕對值
        z_components["oil"] = min(abs(risk.get("oil_ret1", 0.0)) / 2.5, 3.0)
    if "china" in weights:
        # 中國景氣代理(FXI)：中心=0, 1σ=1.5%；下跌越多，對出口/傳產類股風險越高
        z_components["china"] = max(0, -(risk.get("china_ret1", 0.0)) / 1.5)

    raw_score = sum(weights[k] * min(z_components.get(k, 0.0), 3.0) for k in weights)
    continuous_score = round(min(raw_score / 3.0 * 100, 100), 1)

    if continuous_score >= 75:
        regime = "CRASH"
    elif continuous_score >= 50:
        regime = "HIGH_RISK"
    elif continuous_score >= 25:
        regime = "CAUTION"
    else:
        regime = "NORMAL"

    label_map = {
        "vix": ("VIX", risk.get("vix_close", 15.0), "close"),
        "nasdaq": ("NASDAQ", risk.get("nasdaq_ret1", 0.0), "pct"),
        "sox": ("SOX", risk.get("sox_ret1", 0.0), "pct"),
        "us10y": ("US10Y跳升", risk.get("us10y_jump", 0.0), "pct3"),
        "dxy": ("DXY", risk.get("dxy_ret1", 0.0), "pct"),
        "twd": ("台幣匯率", risk.get("twd_ret1", 0.0), "pct"),
        "tw_fin": ("金融指數", risk.get("tw_fin_ret1", 0.0), "pct"),
        "oil": ("油價", risk.get("oil_ret1", 0.0), "pct"),
        "china": ("中國代理(FXI)", risk.get("china_ret1", 0.0), "pct"),
    }
    for k, z in z_components.items():
        if z > 0.5 and k in label_map:
            name, val, fmt = label_map[k]
            if fmt == "close":
                details.append(f"{name}={val:.1f}(z={z:.1f}σ)")
            elif fmt == "pct3":
                details.append(f"{name}{val:+.3f}%(z={z:.1f}σ)")
            else:
                details.append(f"{name}={val:+.1f}%(z={z:.1f}σ)")

    detail_str = f"[{category}類股模型] " + ("、".join(details) if details else "相關風險因子正常")
    return continuous_score, regime, detail_str


def apply_global_risk_override(signal: str, advice: str, pred_5d: float,
                                prob_5d: float, composite_score: float,
                                risk_score: float, risk_regime: str,
                                risk_detail: str) -> tuple:
    """
    V16 升級：全球風險強制覆寫信號 + 動態 Position Sizing。
    
    risk_score 為 0~100 連續分數（V16 量化版）：
      - 懲罰預測報酬：score 越高，預測值折扣越大
      - 動態倉位：根據 regime 自動縮倉
    
    回傳 (signal, advice, pred_5d_adj, composite_score_adj, max_position_size)
    """
    # ── 連續懲罰：risk_score 0~100 → 最大扣 3%（避免過度懲罰） ─
    risk_penalty = -(risk_score / 100) * 3.0
    pred_5d_adj  = pred_5d + risk_penalty
    composite_adj = composite_score + risk_penalty * 2

    # ── 動態倉位上限 ─────────────────────────────────────────────
    if risk_regime == "CRASH":
        max_pos = POSITION_SIZE_CRASH
    elif risk_regime == "HIGH_RISK":
        # HIGH_RISK 內再細分：50~75 分線性縮倉
        t = (risk_score - 50) / 25.0   # 0~1
        max_pos = POSITION_SIZE_HIGHRISK + (1 - t) * (POSITION_SIZE_CAUTION - POSITION_SIZE_HIGHRISK)
        max_pos = round(max(POSITION_SIZE_HIGHRISK, min(max_pos, POSITION_SIZE_CAUTION)), 3)
    elif risk_regime == "CAUTION":
        # CAUTION 內：25~50 分線性縮倉
        t = (risk_score - 25) / 25.0
        max_pos = POSITION_SIZE_BASE + (1 - t) * (POSITION_SIZE_CAUTION - POSITION_SIZE_BASE)
        max_pos = round(max(POSITION_SIZE_CAUTION, min(max_pos, POSITION_SIZE_BASE)), 3)
    else:
        max_pos = POSITION_SIZE_BASE

    max_pos_str = f"{max_pos * 100:.0f}%" if max_pos > 0 else "⚫ 全現金"

    # ── 信號覆寫 ─────────────────────────────────────────────────
    if risk_regime == "CRASH":
        signal = "賣出"
        advice = (f"⚫ 全球崩盤警報｜{risk_detail}｜"
                  f"風險分數:{risk_score:.0f}/100，強制切換【賣出/全面撤退】"
                  f"｜倉位上限:{max_pos_str}｜AI預測(調整後):{pred_5d_adj:.2f}%")

    elif risk_regime == "HIGH_RISK":
        if signal in ["強力買入", "買入", "🔥 黑馬起漲"]:
            signal = "減碼"
            advice = (f"🔴 全球高風險警戒｜{risk_detail}｜"
                      f"風險分數:{risk_score:.0f}/100，強制【減碼】"
                      f"｜倉位上限:{max_pos_str}，技術面偏多但美股衝擊覆寫")
        elif signal == "持平觀望":
            signal = "減碼"
            advice = (f"🔴 全球高風險｜{risk_detail}｜風險分數:{risk_score:.0f}/100，"
                      f"觀望升級【減碼】｜倉位上限:{max_pos_str}")

    elif risk_regime == "CAUTION":
        if signal == "強力買入":
            signal = "買入"
            advice = (f"🟠 全球警戒｜{risk_detail}｜風險分數:{risk_score:.0f}/100，"
                      f"強力買入→【買入】｜倉位上限:{max_pos_str}，縮減資金")
        elif signal in ["買入", "🔥 黑馬起漲"]:
            signal = "持平觀望"
            advice = (f"🟡 全球輕度警戒｜{risk_detail}｜風險分數:{risk_score:.0f}/100，"
                      f"買入→【持平觀望】｜倉位上限:{max_pos_str}，等待美股穩定")
    else:
        # 正常市場：VIX 微高時輕微縮倉提示（不改信號）
        if prob_5d > 65 and risk_score > 10:
            advice = advice + f"｜全球風險:{risk_score:.0f}/100(低)，正常操作"

    return signal, advice, pred_5d_adj, composite_adj, max_pos_str, max_pos


def classify_stock_category(stock_id: str, company: str) -> str:
    """
    V18.2：輕量類股分類（供動態 AI Gate 門檻 / 分業風險模型使用）。
    Google Sheet 股票清單僅有代號與公司名稱，沒有正式產業欄位，因此以：
      1. 代號規則（00開頭多為ETF）
      2. 公司名稱關鍵字（金融/科技/傳產關鍵字表）
    做輔助分類，非精確產業分類。若之後有正式類股欄位，建議直接改讀該欄位取代本函式。
    回傳："科技" / "金融" / "傳產" / "ETF" / "default"
    """
    sid = str(stock_id).strip()
    name = str(company).strip()

    if sid.startswith("00"):
        return "ETF"
    if any(kw in name for kw in FINANCE_KEYWORDS):
        return "金融"
    if any(kw in name for kw in TECH_KEYWORDS):
        return "科技"
    if any(kw in name for kw in TRADITIONAL_KEYWORDS):
        return "傳產"
    return "default"


def classify_sector_fine(stock_id: str, company: str) -> str:
    """
    V18.3：比 classify_stock_category 更細的「主題型」類股分類，供資金輪動 Sector Heat Score 使用。
    比對順序由細到粗：AI → 半導體 → 記憶體 → PCB → 散熱 → 網通 → 航運 → 金融 → 傳產 → ETF → 其他。
    同樣為輕量關鍵字分類，非精確產業分類。
    """
    sid = str(stock_id).strip()
    name = str(company).strip()

    for tag, keywords in SECTOR_FINE_KEYWORDS.items():
        if any(kw in name for kw in keywords):
            return tag
    if sid.startswith("00"):
        return "ETF"
    return "其他"


def compute_stock_momentum(df_price: pd.DataFrame) -> dict:
    """
    V18.3：輕量動能計算（不需完整技術指標/ML流程），供 Sector Heat Score 用的資金輪動偵測。
    結合「近5日價格動能」與「近5日均量 / 近20日均量的量能擴張倍率」，量能擴張代表資金正在進駐。
    """
    n = len(df_price)
    if n < 21:
        return {"ret5": 0.0, "vol_ratio": 1.0, "momentum": 0.0}
    c = df_price["close"]
    v = df_price["volume"]
    lookback = min(SECTOR_HEAT_LOOKBACK_DAYS, n - 1)
    ret5 = float((c.iloc[-1] / c.iloc[-1 - lookback] - 1) * 100) if c.iloc[-1 - lookback] else 0.0
    vol_ratio = float(v.tail(lookback).mean() / (v.tail(20).mean() + 1e-9))
    # 動能 = 70% 價格動能 + 30% 量能擴張加分（量能擴張倍率超過1，代表資金流入該股/該類股）
    momentum = 0.7 * ret5 + 0.3 * (vol_ratio - 1.0) * 20
    return {"ret5": round(ret5, 2), "vol_ratio": round(vol_ratio, 2), "momentum": round(momentum, 2)}


def compute_sector_heat_scores(stocks: list, price_cache: dict) -> tuple:
    """
    V18.3：資金輪動 Sector Heat Score —— 解決「模型只看單股，剛買進資金已撤退到下一個類股」的問題。
    做法：
      1. 對本次監控清單中每一檔股票，計算輕量動能 compute_stock_momentum()（不跑完整ML，避免拖慢速度）。
      2. 依 classify_sector_fine() 分組，取各類股「平均動能」。
      3. 用 z-score 正規化（以本次清單所有類股的平均動能為母體），轉成 0~100 熱度分數（50=中性）。
      4. 若某細分類股本次清單中樣本數 < SECTOR_HEAT_MIN_MEMBERS，視為樣本不足，熱度給中性 50 分。

    回傳 (heat_map: {sector_tag: heat_score}, detail_map: {sector_tag: {"count","avg_ret5","avg_vol_ratio","momentum"}})
    price_cache 會被就地更新（stock_id -> df_price），供主迴圈重複使用，避免對同一檔股票重複下載。
    """
    sector_momentums = {}
    for ticker, stock_id, company in stocks:
        df_price = price_cache.get(stock_id)
        if df_price is None:
            df_price = fetch_yahoo_data(ticker)
            price_cache[stock_id] = df_price
        if df_price is None or df_price.empty or len(df_price) < 21:
            continue
        tag = classify_sector_fine(stock_id, company)
        mom = compute_stock_momentum(df_price)
        sector_momentums.setdefault(tag, []).append(mom)

    detail_map, raw_scores = {}, {}
    for tag, records in sector_momentums.items():
        count = len(records)
        avg_momentum = float(np.mean([r["momentum"] for r in records]))
        avg_ret5 = float(np.mean([r["ret5"] for r in records]))
        avg_vol_ratio = float(np.mean([r["vol_ratio"] for r in records]))
        detail_map[tag] = {"count": count, "avg_ret5": round(avg_ret5, 2),
                            "avg_vol_ratio": round(avg_vol_ratio, 2), "momentum": round(avg_momentum, 2)}
        if count >= SECTOR_HEAT_MIN_MEMBERS:
            raw_scores[tag] = avg_momentum

    heat_map = {}
    if raw_scores:
        vals = list(raw_scores.values())
        mean_m, std_m = float(np.mean(vals)), float(np.std(vals) or 1e-9)
        for tag, m in raw_scores.items():
            z = (m - mean_m) / std_m
            heat_map[tag] = round(min(max(50 + z * 20, 0), 100), 1)
    # 樣本不足或未入選的類股，一律給中性 50 分（避免用不可靠的小樣本強行排名）
    for tag in detail_map:
        heat_map.setdefault(tag, 50.0)

    return heat_map, detail_map


def load_stock_list(gid: str = None, sheet_id: str = None) -> list:
    """從 Google Sheet 指定分頁(gid)載入股票清單。可傳入 gid/sheet_id 覆寫模組預設值。"""
    _sheet_id = sheet_id or SHEET_ID
    _gid = gid if gid is not None else GID
    url = f"https://docs.google.com/spreadsheets/d/{_sheet_id}/export?format=csv&gid={_gid}"
    try: 
        print(f"🔗 正在從 Google Sheet 載入股票清單 (gid={_gid})...")
        df = pd.read_csv(url, header=0)
    except Exception as e:
        print(f"❌ 無法讀取 Google Sheet：{e}")
        return []
    
    stocks = []
    for _, row in df.iterrows():
        ticker = str(row.iloc[0]).strip()
        company = str(row.iloc[1]).strip()
        if not ticker or ticker.lower() == "nan": continue
        stock_id = ticker.replace(".TW", "").replace(".TWO", "")
        stocks.append((ticker, stock_id, company))
    print(f"✅ 成功讀取到 {len(stocks)} 檔股票")
    return stocks


def build_custom_stock_list(codes_text: str) -> list:
    """
    【自選股】依使用者輸入的股號字串組出 stocks 清單，格式與 load_stock_list() 相同：
    [(ticker, stock_id, company), ...]
    輸入範例："2330,2317,0050" 或 "2330.TW,6488.TWO,0050.TW"（可用逗號、頓號、空白、換行分隔）。
    未指定副檔市場時，預設補上 .TW（上市）；OTC(上櫃)個股請自行加上 .TWO。
    公司名稱會嘗試透過 yfinance 自動查詢，查不到則以股號代替。
    """
    if not codes_text or not str(codes_text).strip():
        return []
    raw = re.split(r"[,\uff0c、\s]+", codes_text.strip())
    stocks = []
    seen = set()
    for token in raw:
        token = token.strip()
        if not token:
            continue
        if token.upper().endswith(".TW") or token.upper().endswith(".TWO"):
            ticker = token.upper()
            stock_id = ticker.replace(".TW", "").replace(".TWO", "")
        else:
            stock_id = token
            ticker = f"{token}.TW"
        if stock_id in seen:
            continue
        seen.add(stock_id)

        company = stock_id
        try:
            info = yf.Ticker(ticker).info
            name = info.get("longName") or info.get("shortName")
            if name:
                company = name
        except Exception:
            pass
        stocks.append((ticker, stock_id, company))
    print(f"✅ 自選股清單共 {len(stocks)} 檔：{[s[1] for s in stocks]}")
    return stocks

def fetch_yahoo_data(ticker: str, period: str = PERIOD) -> pd.DataFrame:
    """
    V16：統一使用 date-only index（tz-naive, UTC normalized）確保跨市場對齊。
    """
    try:
        df = yf.download(ticker, period=period, auto_adjust=True, progress=False)
        if df.empty: return pd.DataFrame()
        
        if isinstance(df.columns, pd.MultiIndex): 
            df.columns = df.columns.get_level_values(0)
        df.columns = [c.lower() for c in df.columns]
        
        # V16 嚴格時間軸正規化：統一轉為 tz-naive date（無時區）
        if df.index.tz is not None:
            df.index = df.index.tz_convert("UTC").tz_localize(None)
        # 強制轉為 date-only（去除 intraday 時間部分）
        df.index = pd.to_datetime(df.index.date)
            
        df = df[~df.index.duplicated(keep='last')].sort_index()
        # 移除全欄 NaN 的行（常見於節假日填充）
        df.dropna(how='all', inplace=True)
        return df
    except Exception as e:
        print(f"  ❌ 下載 {ticker} 價量資料失敗: {e}")
        return pd.DataFrame()

def fetch_chip_data(stock_id: str, dl) -> pd.DataFrame:
    start_date = (datetime.now() - pd.DateOffset(years=2)).strftime('%Y-%m-%d')
    df_inst, df_margin = pd.DataFrame(), pd.DataFrame()
    
    # 外資與投信買賣超
    for retry in range(3):
        try:
            time.sleep(1.5 + retry * 2)
            df_inst = dl.taiwan_stock_institutional_investors(stock_id=stock_id, start_date=start_date)
            break
        except Exception as e:
            if "upper limit" in str(e).lower():
                print("    ⚠️ API 額度上限，略過外資投信資料。")
                break
            time.sleep(5)
            
    # 信用交易融資券餘額
    for retry in range(3):
        try:
            time.sleep(1.5 + retry * 2)
            df_margin = dl.taiwan_stock_margin_purchase_short_sale(stock_id=stock_id, start_date=start_date)
            break
        except Exception as e:
            if "upper limit" in str(e).lower():
                print("    ⚠️ API 額度上限，略過融資券資料。")
                break
            time.sleep(5)

    chip_df = pd.DataFrame()

    if isinstance(df_inst, pd.DataFrame) and not df_inst.empty:
        df_inst['net'] = df_inst['buy'] - df_inst['sell']
        df_inst_pivot = df_inst.pivot_table(index='date', columns='name', values='net', aggfunc='sum').fillna(0)
        cols = {}
        if 'Foreign_Investor' in df_inst_pivot.columns: cols['Foreign_Investor'] = 'Foreign_Net'
        if 'Investment_Trust' in df_inst_pivot.columns: cols['Investment_Trust'] = 'Trust_Net'
        df_inst_pivot.rename(columns=cols, inplace=True)
        chip_df = df_inst_pivot.copy()

    if isinstance(df_margin, pd.DataFrame) and not df_margin.empty:
        cols_to_keep = ['date']
        rename_dict = {}
        if 'MarginPurchaseTodayBalance' in df_margin.columns:
            cols_to_keep.append('MarginPurchaseTodayBalance')
            rename_dict['MarginPurchaseTodayBalance'] = 'Margin_Bal'
        if 'ShortSaleTodayBalance' in df_margin.columns:
            cols_to_keep.append('ShortSaleTodayBalance')
            rename_dict['ShortSaleTodayBalance'] = 'Short_Bal'

        if len(cols_to_keep) > 1:
            df_margin_clean = df_margin[cols_to_keep].copy()
            df_margin_clean.rename(columns=rename_dict, inplace=True)
            df_margin_clean = df_margin_clean.drop_duplicates(subset=['date'])
            df_margin_clean.set_index('date', inplace=True)
            if chip_df.empty: chip_df = df_margin_clean
            else: chip_df = chip_df.join(df_margin_clean, how='outer')

    if not chip_df.empty:
        chip_df.index = pd.to_datetime(chip_df.index)
        if chip_df.index.tz is not None:
            chip_df.index = chip_df.index.tz_localize(None)
        chip_df = chip_df[~chip_df.index.duplicated(keep='last')].sort_index()
        
    return chip_df

def process_macro_data(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty: return pd.DataFrame()
        
    c = df["close"]
    df["macro_sma20"]      = SMAIndicator(c, 20).sma_indicator()
    df["macro_sma60"]      = SMAIndicator(c, 60).sma_indicator()
    df["macro_bias20"]     = (c - df["macro_sma20"]) / (df["macro_sma20"] + 1e-9) * 100
    macd_obj              = MACD(c)
    df["macro_macd_hist"] = macd_obj.macd_diff()
    df["macro_ret20"]      = c.pct_change(20) * 100
    df["macro_ret1d"]      = c.pct_change(1) * 100

    macro_cols = ["macro_sma20", "macro_sma60", "macro_bias20", "macro_macd_hist",
                  "macro_ret20", "macro_ret1d", "close"]

    if "ADL" in df.columns:
        df["macro_ADL_trend"] = df["ADL"] - df["ADL"].rolling(20).mean()
        macro_cols.append("macro_ADL_trend")
    if "Foreign_OI" in df.columns:
        df["macro_Foreign_OI_trend"] = df["Foreign_OI"].diff(5)
        macro_cols.append("macro_Foreign_OI_trend")

    def classify_regime(row):
        close = row["close"]
        sma20 = row["macro_sma20"]
        sma60 = row["macro_sma60"]
        macdh = row["macro_macd_hist"]

        if pd.isna(sma20) or pd.isna(sma60): return "低波動盤整"
        bias = ((close - sma20) / (sma20 + 1e-9)) * 100

        if close < sma60 * 0.9 and macdh < 0: return "崩盤"
        if close < sma60 and macdh < 0: return "主跌段"
        if close < sma20 and macdh > 0: return "空頭反彈"
        if close > sma20 > sma60 and bias > 5 and macdh > 0: return "強趨勢多頭"
        if close > sma20 > sma60: return "緩漲盤"
        if abs(bias) > 3: return "高波動震盪"
        return "低波動盤整"

    df["macro_regime"] = df.apply(classify_regime, axis=1)
    
    regime_score = {
        "強趨勢多頭": 3, "緩漲盤": 2, "高波動震盪": 0, "低波動盤整": -1,
        "空頭反彈": -2, "主跌段": -3, "崩盤": -4
    }
    df["macro_regime_score"] = df["macro_regime"].map(regime_score)
    macro_cols.extend(["macro_regime", "macro_regime_score"])

    return df[macro_cols].rename(columns={"close": "macro_close"})

def compute_indicators(df: pd.DataFrame) -> pd.DataFrame:
    c, h, l, v, o = df["close"], df["high"], df["low"], df["volume"], df["open"]

    for w in [6, 14, 21]: df[f"RSI_{w}"] = RSIIndicator(c, w).rsi()

    macd = MACD(c)
    df["MACD_line"], df["MACD_signal"], df["MACD_hist"] = macd.macd(), macd.macd_signal(), macd.macd_diff()
    df["MACD_hist_slope"] = df["MACD_hist"] - df["MACD_hist"].shift(1)

    for w in [10, 20]:
        bb = BollingerBands(c, w)
        df[f"BB_upper_{w}"], df[f"BB_lower_{w}"], df[f"BB_mid_{w}"], df[f"BB_pct_{w}"] = \
            bb.bollinger_hband(), bb.bollinger_lband(), bb.bollinger_mavg(), bb.bollinger_pband()

    # ================= V18：BB 模組（新增，僅追加特徵，不變更既有模型流程） =================
    # BB Width：布林通道寬度百分比（以 20 日通道為主，波段訊號較穩定）
    df["BB_width_20"] = (df["BB_upper_20"] - df["BB_lower_20"]) / (df["BB_mid_20"] + 1e-9) * 100
    # BB Width Change：5 日通道寬度變化（負值代表持續收縮）
    df["BB_width_chg_5"] = df["BB_width_20"].diff(5)
    # BB Width ZScore：以近 60 日通道寬度分布做標準化
    _bbw_mean60 = df["BB_width_20"].rolling(60).mean()
    _bbw_std60  = df["BB_width_20"].rolling(60).std()
    df["BB_width_zscore"] = (df["BB_width_20"] - _bbw_mean60) / (_bbw_std60 + 1e-9)
    # BB Squeeze：通道寬度顯著低於近期常態（zscore < -1）視為擠壓成立
    df["BB_squeeze"] = (df["BB_width_zscore"] < -1.0).astype(int)
    # BB Width 歷史百分位（近 120 日）：供 HorseFinder 判斷「是否低於歷史10%分位」
    df["BB_width_pct_rank"] = df["BB_width_20"].rolling(120, min_periods=60).apply(
        lambda x: float((x <= x.iloc[-1]).mean()), raw=False
    )
    # 強化 BB %B：標記突破上下軌的極端狀態（1=突破上軌，-1=跌破下軌，0=通道內）
    df["BB_pct_extreme"] = np.where(df["BB_pct_20"] > 1.0, 1, np.where(df["BB_pct_20"] < 0.0, -1, 0))
    # BB Confidence Bonus：近 3 日內曾發生擠壓、且目前 %B 已站上通道上緣 0.8，代表擠壓後轉強，加分納入信心等級評估
    df["BB_confidence_bonus"] = (
        (df["BB_squeeze"].shift(1).rolling(3).max().fillna(0) == 1) & (df["BB_pct_20"] > 0.8)
    ).astype(int) * 10

    for w in [5, 10, 20, 60]:
        df[f"SMA_{w}"], df[f"EMA_{w}"] = SMAIndicator(c, w).sma_indicator(), EMAIndicator(c, w).ema_indicator()
    for w in [5, 20, 60]: df[f"Bias_SMA_{w}"] = (c - df[f"SMA_{w}"]) / (df[f"SMA_{w}"] + 1e-9) * 100

    adx = ADXIndicator(h, l, c, 14)
    df["ADX"] = adx.adx()
    df["DI_pos"] = adx.adx_pos()
    df["DI_neg"] = adx.adx_neg()

    # 進場/波段關鍵動能因子
    df["Low_5"] = l.rolling(5).min()
    df["EMA_10"] = EMAIndicator(c, 10).ema_indicator()

    ema20 = EMAIndicator(c, 20).ema_indicator()
    ema60 = EMAIndicator(c, 60).ema_indicator()
    df["EMA20_slope"] = ema20.diff(5)
    df["EMA60_slope"] = ema60.diff(5)

    df["Breakout_20"] = (c > h.rolling(20).max().shift(1)).astype(int)
    df["Breakdown_20"] = (c < l.rolling(20).min().shift(1)).astype(int)
    df["Volume_Break"] = (v > v.rolling(20).mean() * 2).astype(int)

    # V14 高級遺失因子補回
    vol_10 = c.pct_change().rolling(10).std()
    vol_mean_20 = vol_10.rolling(20).mean()
    vol_std_20 = vol_10.rolling(20).std()
    df["Volatility_Squeeze"] = (vol_10 - vol_mean_20) / (vol_std_20 + 1e-9)

    df["Momentum_Strength"] = df["EMA20_slope"] * df["ADX"]
    df["Trend_Strength"] = df["DI_pos"] - df["DI_neg"]
    
    df["Volume_Dryup"] = (v < v.rolling(20).mean() * 0.5).astype(int)
    df["Trend_Acceleration"] = df["EMA20_slope"] - df["EMA20_slope"].shift(5)
    
    rsi_14 = RSIIndicator(c, 14).rsi()
    df["RSI_divergence"] = ((c >= h.rolling(20).max()) & (rsi_14 < rsi_14.rolling(20).max())).astype(int)

    stoch = StochasticOscillator(h, l, c)
    df["KD_K"], df["KD_D"] = stoch.stoch(), stoch.stoch_signal()

    df["High_20"], df["Low_20"] = h.rolling(20).max(), l.rolling(20).min()
    df["Dist_High_20"] = (c - df["High_20"]) / (df["High_20"] + 1e-9) * 100
    df["Dist_Low_20"]  = (c - df["Low_20"])  / (df["Low_20"] + 1e-9)  * 100

    for w in [7, 14]: df[f"ATR_{w}"] = AverageTrueRange(h, l, c, w).average_true_range()

    df["OBV"]       = OnBalanceVolumeIndicator(c, v).on_balance_volume()
    df["VOL_ratio"] = v / (v.rolling(20).mean() + 1e-9)
    df["VOL_ratio_5"]  = v / (v.rolling(5).mean() + 1e-9)
    df["VOL_ratio_60"] = v / (v.rolling(60).mean() + 1e-9)

    # ★ V18.2 專家建議：成交值(金額)爆量倍率，常比單純股數爆量更早反映法人資金進場
    df["Turnover"] = c * v
    df["Turnover_Ratio_20"] = df["Turnover"] / (df["Turnover"].rolling(20).mean() + 1e-9)

    for lag in [1, 2, 3, 5, 10, 20]: df[f"Ret_{lag}d"] = c.pct_change(lag) * 100
    
    df["HL_pct"] = (h - l) / (c + 1e-9) * 100
    df["OC_pct"] = (c - o) / (o + 1e-9) * 100
    df["HV_10"]  = c.pct_change().rolling(10).std() * np.sqrt(252) * 100
    df["Mom_5"]  = c - c.shift(5)
    df["Gap_pct"] = (o - c.shift(1)) / (c.shift(1) + 1e-9) * 100
    
    if 'macro_ret1d' in df.columns:
        stock_ret1d = c.pct_change(1) * 100
        df['Beta_20'] = stock_ret1d.rolling(20).cov(df['macro_ret1d']) / (df['macro_ret1d'].rolling(20).var() + 1e-9)
        if 'macro_ret20' in df.columns:
            df['Alpha_20'] = df['Ret_20d'] - df['Beta_20'] * df['macro_ret20']
            df['relative_strength_20d'] = df['Ret_20d'] - df['macro_ret20']

    if 'Foreign_Net' in df.columns:
        df['Foreign_Net_5d']  = df['Foreign_Net'].rolling(5).sum()
        df['Foreign_Vol_Ratio'] = df['Foreign_Net'] / (df['volume'] + 1) * 100
        df['foreign_buy_streak'] = df['Foreign_Net'].gt(0).groupby((df['Foreign_Net'].gt(0) != df['Foreign_Net'].gt(0).shift()).cumsum()).cumsum().where(df['Foreign_Net'].gt(0), 0)
        
    if 'Trust_Net' in df.columns:
        df['Trust_Net_5d']    = df['Trust_Net'].rolling(5).sum()
        df['Trust_Vol_Ratio'] = df['Trust_Net'] / (df['volume'] + 1) * 100
        df['trust_buy_streak'] = df['Trust_Net'].gt(0).groupby((df['Trust_Net'].gt(0) != df['Trust_Net'].gt(0).shift()).cumsum()).cumsum().where(df['Trust_Net'].gt(0), 0)
        
    if 'Margin_Bal' in df.columns: df['Margin_Change_5d'] = df['Margin_Bal'].pct_change(5) * 100
    if 'Short_Bal' in df.columns:  df['Short_Change_5d']  = df['Short_Bal'].pct_change(5) * 100
    if 'Short_Bal' in df.columns and 'Margin_Bal' in df.columns:
        df['Short_Margin_Ratio'] = (df['Short_Bal'] / (df['Margin_Bal'] + 1)) * 100

    return df

def define_target(df: pd.DataFrame) -> pd.DataFrame:
    for days in HORIZONS:
        df[f"target_{days}d"] = df["close"].shift(-days) / df["close"] * 100 - 100
        df[f"target_up_{days}d"] = (df[f"target_{days}d"] > CLS_THRESHOLD).astype(int)
        future_min = df["low"].shift(-days).rolling(days).min()
        df[f"future_drawdown_{days}d"] = future_min / df["close"] * 100 - 100
    
    df["target_cls_5d"] = df["target_up_5d"]
    return df

def feature_selection(df: pd.DataFrame):
    target_cols = [f"target_{d}d" for d in HORIZONS] + [f"target_up_{d}d" for d in HORIZONS] + [f"future_drawdown_{d}d" for d in HORIZONS] + ["target_cls_5d"]
    exclude = {"open", "high", "low", "close", "volume", "High_20", "Low_20", "Margin_Bal", "Short_Bal",
               "macro_close", "macro_sma20", "macro_sma60", "macro_regime", "macro_ret1d"}
    exclude.update(target_cols)
    feature_cols = [col for col in df.columns if col not in exclude]

    df_clean = (df[feature_cols + ["target_5d"]].replace([np.inf, -np.inf], np.nan).dropna())
    for col in feature_cols:
        if col in df_clean.columns:
            mean, std = df_clean[col].mean(), df_clean[col].std()
            if std > 0: df_clean = df_clean[np.abs(df_clean[col] - mean) <= 5 * std]

    if len(df_clean) < 100: return None, None

    train_cut = int(len(df_clean) * 0.80)
    df_fs = df_clean.iloc[:train_cut]
    X_fs, y_fs = df_fs[feature_cols], df_fs["target_5d"]
    
    corr_matrix = X_fs.corr().abs()
    upper = corr_matrix.where(np.triu(np.ones(corr_matrix.shape), k=1).astype(bool))
    drop_cols = [col for col in upper.columns if any(upper[col] > 0.92)]
    feature_cols = [c for c in feature_cols if c not in drop_cols]
    
    X_fs = X_fs[feature_cols]

    corr_records = []
    for col in feature_cols:
        try: r_p = pearsonr(X_fs[col], y_fs)[0]
        except: r_p = 0.0
        if pd.isna(r_p) or np.isnan(r_p): r_p = 0.0

        try: r_s = spearmanr(X_fs[col], y_fs)[0]
        except: r_s = 0.0
        if pd.isna(r_s) or np.isnan(r_s): r_s = 0.0

        corr_records.append({"feature": col, "pearson": round(r_p, 4), "spearman": round(r_s, 4), "abs_pearson": abs(r_p)})
    corr_df = pd.DataFrame(corr_records)

    try:
        mi_scores = mutual_info_regression(X_fs, y_fs, random_state=42)
        mi_max = mi_scores.max() if mi_scores.max() > 0 else 1.0
        mi_norm = mi_scores / mi_max
    except: mi_norm = np.zeros(len(feature_cols))
    mi_df = pd.DataFrame({"feature": feature_cols, "mutual_info": mi_norm})

    tscv = TimeSeriesSplit(n_splits=5)
    xgb_imp, lgb_imp = np.zeros(len(feature_cols)), np.zeros(len(feature_cols))
    
    X_train_only = df_clean.iloc[:train_cut][feature_cols]
    y_train_only = df_clean.iloc[:train_cut]["target_5d"]

    eval_params = {'n_estimators': 50, 'max_depth': 4, 'learning_rate': 0.05, 'random_state': 42, 'n_jobs': 4}

    for tr_idx, te_idx in tscv.split(X_train_only):
        y_tr_clipped = np.clip(y_train_only.iloc[tr_idx], TARGET_CLIP_MIN, TARGET_CLIP_MAX)
        X_tr_np = X_train_only.iloc[tr_idx].astype(np.float32).values
        y_tr_np = np.array(y_tr_clipped, dtype=np.float32)
        
        m_x = xgb.XGBRegressor(**eval_params)
        m_x.fit(X_tr_np, y_tr_np)
        xgb_imp += m_x.feature_importances_
        m_l = lgb.LGBMRegressor(n_estimators=50, max_depth=4, learning_rate=0.05, random_state=42, verbose=-1, n_jobs=1)
        m_l.fit(X_tr_np, y_tr_np)
        lgb_imp += m_l.feature_importances_

    xgb_imp, lgb_imp = xgb_imp / 5, lgb_imp / 5
    imp_df = pd.DataFrame({"feature": feature_cols, "xgb_importance": xgb_imp, "lgb_importance": lgb_imp})
    for col in ["xgb_importance", "lgb_importance"]:
        if imp_df[col].max() > 0: imp_df[col] /= imp_df[col].max()

    # ================= V18：SHAP 全域特徵重要度（新增，未安裝 shap 時自動略過） =================
    shap_importance = np.zeros(len(feature_cols))
    if HAS_SHAP:
        try:
            model_shap = xgb.XGBRegressor(**eval_params)
            y_tr_full_clipped = np.clip(y_train_only, TARGET_CLIP_MIN, TARGET_CLIP_MAX)
            model_shap.fit(X_train_only.astype(np.float32).values, np.array(y_tr_full_clipped, dtype=np.float32))
            explainer = shap.TreeExplainer(model_shap)
            sample_X = X_train_only.tail(200).astype(np.float32).values if len(X_train_only) > 200 else X_train_only.astype(np.float32).values
            shap_values = explainer.shap_values(sample_X)
            shap_importance = np.abs(shap_values).mean(axis=0)
            if shap_importance.max() > 0: shap_importance = shap_importance / shap_importance.max()
        except Exception as e:
            print(f"  ⚠️ SHAP 全域重要度計算失敗，略過：{e}")
            shap_importance = np.zeros(len(feature_cols))
    imp_df["shap_importance"] = shap_importance

    imp_df["avg_importance"] = (imp_df["xgb_importance"] + imp_df["lgb_importance"]) / 2
    result = imp_df.merge(corr_df, on="feature").merge(mi_df, on="feature")
    # V18：final_score 納入 SHAP 重要度（權重總和維持 1.0）
    result["final_score"] = (result["avg_importance"] * 0.40 + result["shap_importance"] * 0.15
                              + result["mutual_info"] * 0.25 + result["abs_pearson"] * 0.20)
    result = result.sort_values("final_score", ascending=False).reset_index(drop=True)
    result["rank"] = result.index + 1

    effective_pool = result[result["avg_importance"] >= MIN_IMPORTANCE].copy()

    # ── V16：Feature 數量上限控制 ─────────────────────────────
    total_candidates = len(effective_pool)
    final_top_n = min(TOP_N_FEATURES, MAX_FEATURES)

    if total_candidates > MAX_FEATURES:
        print(f"  ⚠️  [V16 Feature Guard] 有效因子共 {total_candidates} 個，"
              f"超過上限 MAX_FEATURES={MAX_FEATURES}，強制截取前 {final_top_n} 個以防 overfitting")
    elif total_candidates > TOP_N_FEATURES:
        print(f"  ℹ️  有效因子 {total_candidates} 個，取前 {final_top_n} 個核心因子")
    else:
        print(f"  ✅ 有效因子 {total_candidates} 個（在上限內），全數採用")

    return result, effective_pool.head(final_top_n).copy()

def optimize_hyperparameters_random(X, y):
    """V17 版超參數搜尋：RandomizedSearchCV（作為 Optuna 未安裝時的自動退回方案）"""
    param_dist = {
        'max_depth': [3, 4, 5],
        'learning_rate': [0.01, 0.03, 0.05],
        'subsample': [0.7, 0.8, 0.9],
        'reg_alpha': [0.0, 0.5, 1.0],      # L1 Regularization
        'reg_lambda': [1.0, 3.0, 5.0]      # L2 Regularization
    }
    tscv = TimeSeriesSplit(n_splits=3)
    model = xgb.XGBRegressor(n_estimators=50, random_state=42, n_jobs=2)
    
    X_tune = X.iloc[-500:] if len(X) > 500 else X
    y_tune = y.iloc[-500:] if len(y) > 500 else y
    
    search = RandomizedSearchCV(model, param_distributions=param_dist, n_iter=8, cv=tscv, random_state=42, n_jobs=1)
    search.fit(X_tune.astype(np.float32).values, y_tune.astype(np.float32).values)
    
    return search.best_params_


def optimize_hyperparameters_optuna(X, y, n_trials: int = 20) -> dict:
    """V18：Optuna 貝葉斯優化調參，以 TimeSeriesSplit 內層 CV 的 MSE 為目標函數。"""
    tscv = TimeSeriesSplit(n_splits=3)
    X_tune = X.iloc[-500:] if len(X) > 500 else X
    y_tune = y.iloc[-500:] if len(y) > 500 else y
    X_vals = X_tune.astype(np.float32).values
    y_vals = np.clip(y_tune.astype(np.float32).values, TARGET_CLIP_MIN, TARGET_CLIP_MAX)

    def objective(trial):
        params = {
            'max_depth': trial.suggest_int('max_depth', 3, 6),
            'learning_rate': trial.suggest_float('learning_rate', 0.01, 0.08, log=True),
            'subsample': trial.suggest_float('subsample', 0.6, 0.95),
            'reg_alpha': trial.suggest_float('reg_alpha', 0.0, 2.0),
            'reg_lambda': trial.suggest_float('reg_lambda', 1.0, 6.0),
        }
        fold_scores = []
        for tr_idx, te_idx in tscv.split(X_vals):
            model = xgb.XGBRegressor(n_estimators=80, random_state=42, n_jobs=2, **params)
            model.fit(X_vals[tr_idx], y_vals[tr_idx])
            pred = model.predict(X_vals[te_idx])
            fold_scores.append(mean_squared_error(y_vals[te_idx], pred))
        return float(np.mean(fold_scores))

    study = optuna.create_study(direction="minimize", sampler=optuna.samplers.TPESampler(seed=42))
    study.optimize(objective, n_trials=n_trials, show_progress_bar=False)
    return study.best_params


def optimize_hyperparameters(X, y):
    """
    V18：優先使用 Optuna 貝葉斯調參以取得更佳超參數；
    若未安裝 optuna 或執行過程發生例外，自動退回 V17 的 RandomizedSearchCV（架構完整保留）。
    """
    if HAS_OPTUNA:
        try:
            return optimize_hyperparameters_optuna(X, y, n_trials=20)
        except Exception as e:
            print(f"  ⚠️ Optuna 調參失敗，改用 V17 RandomizedSearchCV：{e}")
    return optimize_hyperparameters_random(X, y)

def get_regime_label(regime: str) -> str:
    """將大盤環境字串對齊高辨識度 Emoji 標籤"""
    mapping = {"強趨勢多頭": "🟢 強趨勢多頭", "緩漲盤": "🔵 緩漲盤", "高波動震盪": "🟡 高波動震盪", 
               "低波動盤整": "⚪ 低波動盤整", "空頭反彈": "🟠 空頭反彈", "主跌段": "🔴 主跌段", "崩盤": "⚫ 崩盤"}
    return mapping.get(regime, "⚪ 低波動盤整")

# ==========================================
# 🐎 V18：HorseFinder 獨立黑馬評分機制
# ==========================================
def compute_horse_score(df: pd.DataFrame) -> dict:
    """
    V18 HorseFinder：獨立於 AI 模型分數之外的黑馬候選股評分。
    此分數「不會」寫回模型特徵、也不會影響 AI 預測分數（composite_score），
    純粹作為 Excel 報表中「黑馬候選股」識別與排行使用。

    配分表（V18.2 更新，加入資金流因子）：
      BB Squeeze                       +20
      BB Width < 歷史10%分位            +15
      成交量(股數)突破 (>=2倍均量)       +15
      成交值/20日均成交值 > 2.5倍(資金流) +15   ← V18.2新增，比技術指標更早反映法人進場
      20日新高突破                      +15
      MACD 黃金交叉 (近3日內)           +10
      ADX > 25                         +10
      法人連買 (外資或投信連買>=3日)     +10
    HorseScore >= 70 → 准黑馬｜>= 80 → 強黑馬｜>= 90 → 爆發黑馬
    """
    w = HORSE_SCORE_WEIGHTS
    breakdown = {k: 0 for k in w}
    try:
        if len(df) < 5:
            return {"horse_score": 0, "horse_flag": False, "breakdown": breakdown}

        last = df.iloc[-1]

        if int(last.get("BB_squeeze", 0) or 0) == 1:
            breakdown["bb_squeeze"] = w["bb_squeeze"]

        bb_width_pctrank = last.get("BB_width_pct_rank", np.nan)
        if pd.notna(bb_width_pctrank) and bb_width_pctrank <= 0.10:
            breakdown["bb_width_low10pct"] = w["bb_width_low10pct"]

        if int(last.get("Volume_Break", 0) or 0) == 1:
            breakdown["volume_breakout"] = w["volume_breakout"]

        turnover_ratio = last.get("Turnover_Ratio_20", np.nan)
        if pd.notna(turnover_ratio) and turnover_ratio > VALUE_SURGE_RATIO_THRESHOLD:
            breakdown["value_surge_ratio"] = w["value_surge_ratio"]

        if int(last.get("Breakout_20", 0) or 0) == 1:
            breakdown["new_high_20d"] = w["new_high_20d"]

        if "MACD_hist" in df.columns:
            macd_recent = df["MACD_hist"].tail(3).dropna()
            if len(macd_recent) >= 2 and (macd_recent.iloc[:-1] <= 0).any() and macd_recent.iloc[-1] > 0:
                breakdown["macd_golden_cross"] = w["macd_golden_cross"]

        adx_val = last.get("ADX", np.nan)
        if pd.notna(adx_val) and adx_val > 25:
            breakdown["adx_strong"] = w["adx_strong"]

        foreign_streak = last.get("foreign_buy_streak", 0) if "foreign_buy_streak" in df.columns else 0
        trust_streak   = last.get("trust_buy_streak", 0) if "trust_buy_streak" in df.columns else 0
        inst_streak = max(float(foreign_streak or 0), float(trust_streak or 0))
        if inst_streak >= 3:
            breakdown["institution_streak"] = w["institution_streak"]

        horse_score = int(sum(breakdown.values()))
        star_info = horse_score_to_stars(horse_score)
        return {"horse_score": horse_score, "horse_flag": horse_score >= HORSE_SCORE_THRESHOLD,
                "horse_stars": star_info["stars"], "horse_tier": star_info["tier"], "breakdown": breakdown}
    except Exception:
        star_info = horse_score_to_stars(0)
        return {"horse_score": 0, "horse_flag": False, "horse_stars": star_info["stars"], "horse_tier": star_info["tier"], "breakdown": breakdown}


def horse_score_to_stars(horse_score: int) -> dict:
    """V18.1：Horse Score 改為 5 級星等，取代單純 YES/NO，資訊量更連續、更專業。"""
    for threshold, stars, tier in HORSE_STAR_TIERS:
        if horse_score >= threshold:
            return {"stars": stars, "tier": tier}
    return {"stars": "★☆☆☆☆", "tier": "極弱"}


def compute_trend_score(pos_count: int, latest_adx: float) -> float:
    """
    V18.2：把「多週期方向一致性(pos_count 0~4)」與「ADX 趨勢強度」合成一個 0~100 的趨勢分數，
    供動態 AI Gate 門檻判斷「超強趨勢股」使用（趨勢分數與 Sharpe 同時偏高時，AI門檻可再放寬）。
    """
    consistency_part = (pos_count / 4.0) * 60          # 方向一致性最多貢獻 60 分
    adx_part = min(max(latest_adx, 0.0), 50.0) / 50.0 * 40   # ADX 最多貢獻 40 分（ADX>=50視為封頂）
    return round(consistency_part + adx_part, 1)


def classify_market_mode(latest_regime: str, vix_close: float = None) -> str:
    """
    V18.2b：把既有 macro_regime（強趨勢多頭/緩漲盤/高波動震盪/低波動盤整/空頭反彈/主跌段/崩盤）
    收斂成 4 種 MARKET_MODE，供全站策略切換使用：
      BULL     ← 強趨勢多頭 / 緩漲盤
      SIDEWAYS ← 高波動震盪 / 低波動盤整
      BEAR     ← 空頭反彈 / 主跌段
      PANIC    ← 崩盤，或 VIX 恐慌指數超過 VIX_PANIC_THRESHOLD（雙重確認，任一觸發即視為恐慌）
    """
    if latest_regime == "崩盤":
        return "PANIC"
    if vix_close is not None and vix_close >= VIX_PANIC_THRESHOLD:
        return "PANIC"
    return MARKET_MODE_MAP.get(latest_regime, "SIDEWAYS")


def apply_market_mode_filter(signal: str, advice: str, market_mode: str, stock_category: str,
                              latest_rsi: float = 50.0) -> dict:
    """
    V18.2b：市場模式風格過濾器 —— 避免「模型在空頭還硬做突破股」這類風格與市場不匹配的情況。
    邏輯：
      - PANIC：現金為王，任何非防守訊號一律降級為「持平觀望」（減碼/賣出不受影響）。
      - BEAR：只有屬於該模式偏好類股（目前=金融）才維持原強度買進訊號；
              非偏好類股若 RSI 未達超賣(<30)，視為「逆勢追多」風格不符，降級一級並提示；
              若 RSI 已超賣，視為「逆勢低接」候選，予以保留但註記。
      - SIDEWAYS：偏好均值回歸/RSI低接，不追突破；非偏好類股的「強力買入」降級為「買入」並提示宜區間操作。
      - BULL：偏好突破動能，不做額外限制（維持原訊號）。
    """
    playbook = MARKET_MODE_PLAYBOOK.get(market_mode, MARKET_MODE_PLAYBOOK["SIDEWAYS"])
    favored = playbook["favored_categories"]
    label = playbook["label"]
    style_note = playbook["favored_style"]

    if market_mode == "PANIC":
        if signal in ["強力買入", "買入", "持平觀望", "🔥 黑馬起漲", "🚀 搶先布局(Horse Override)"]:
            new_signal = "持平觀望"
            new_advice = f"{label} MARKET_MODE=PANIC｜市場恐慌，現金為王，全面降級為觀望。原訊號「{signal}」。原建議：{advice}"
            return {"signal": new_signal, "advice": new_advice, "mode_adjusted": True}
        return {"signal": signal, "advice": advice, "mode_adjusted": False}

    if market_mode == "BEAR" and stock_category not in favored and signal in ["強力買入", "買入", "🔥 黑馬起漲"]:
        if latest_rsi is not None and latest_rsi < MARKET_MODE_RSI_OVERSOLD:
            new_advice = f"{label} MARKET_MODE=BEAR｜{stock_category}非偏好類股，但RSI={latest_rsi:.1f}已達超賣，視為逆勢低接候選(非順勢突破)，保留訊號但請留意風控。原建議：{advice}"
            return {"signal": signal, "advice": new_advice, "mode_adjusted": True}
        downgrade_map = {"強力買入": "買入", "買入": "持平觀望", "🔥 黑馬起漲": "持平觀望"}
        new_signal = downgrade_map.get(signal, signal)
        new_advice = f"{label} MARKET_MODE=BEAR｜空頭市場偏好「{style_note}」，{stock_category}股突破訊號風格不match，降級為「{new_signal}」。原訊號：{signal}｜原建議：{advice}"
        return {"signal": new_signal, "advice": new_advice, "mode_adjusted": True}

    if market_mode == "SIDEWAYS" and stock_category not in favored and signal == "強力買入":
        new_signal = "買入"
        new_advice = f"{label} MARKET_MODE=SIDEWAYS｜盤整盤偏好「{style_note}」，不宜追高強力買入訊號，降級為「買入」並建議區間操作。原建議：{advice}"
        return {"signal": new_signal, "advice": new_advice, "mode_adjusted": True}

    return {"signal": signal, "advice": advice, "mode_adjusted": False}


def compute_dynamic_ai_gate_threshold(category: str, trend_score: float, sharpe: float,
                                       market_mode: str = None) -> float:
    """
    V18.2：動態 AI Decision Gate 門檻。
    專家建議：固定 50% 對台股很多波段股(尤其低波動金融股)太硬，48~55%的AI機率其實已經足夠賺，
    結果常常「趨勢對了，但AI信心不夠」而錯過主升段。改為：
      1. 依類股別給出基礎門檻（科技55% / ETF、傳產50% / 金融45%）。
      2. 若同時符合「高趨勢分數」與「高歷史Sharpe」（超強趨勢股），門檻再下修 5 個百分點。
      3. 【V18.2b新增】依 MARKET_MODE 再微調：多頭放寬(-3)／盤整收緊(+3)／空頭收緊(+5)／恐慌等同鎖死(+100)。
      4. 無論如何下修，門檻不低於 AI_GATE_THRESHOLD_FLOOR（43%）的保護下限（恐慌模式除外，恐慌本就該鎖死）。
    """
    base_threshold = AI_GATE_THRESHOLD_BY_CATEGORY.get(category, AI_GATE_THRESHOLD_BY_CATEGORY["default"])

    cfg = AI_GATE_TREND_SHARPE_DISCOUNT
    if trend_score >= cfg["trend_score_min"] and sharpe >= cfg["sharpe_min"]:
        base_threshold -= cfg["discount"]

    if market_mode is not None:
        base_threshold += MARKET_MODE_PLAYBOOK.get(market_mode, {}).get("ai_gate_delta", 0)

    if market_mode == "PANIC":
        return base_threshold   # 恐慌模式門檻刻意超過100%，等同鎖死，不受下限保護
    return max(base_threshold, AI_GATE_THRESHOLD_FLOOR)


def apply_ai_decision_gate(signal: str, advice: str, prob_5d: float, horse_result: dict,
                            ai_gate_threshold: float = None, category: str = "default") -> dict:
    """
    V18.1/V18.2：AI Decision Gate（決策層守門機制）
    問題：AI上漲機率偏低時，系統可能因為其他分數（趨勢一致性、技術面）仍給出「買入」，
         造成「AI沒信心，卻建議買入」的過度積極矛盾。

    V18.2 更新重點：
      - 門檻不再固定 50%，改由外部傳入依「類股別+趨勢+Sharpe」動態計算的 ai_gate_threshold
        （若未傳入則 fallback 回 AI_GATE_LOW_PROB=50%，向下相容）。
      - HorseScore Override 不再單一鎖 90 分，改為階梯式（AI_GATE_OVERRIDE_TIERS）：
        90分(爆發黑馬):AI機率>=30%可覆寫／80分(強黑馬):AI機率>=40%可覆寫／70分(准黑馬):AI機率>=45%可覆寫。
      - 若原始訊號本來就是「減碼」或「賣出」（風險控管訊號），Gate 不干預，維持原樣。
    """
    threshold = ai_gate_threshold if ai_gate_threshold is not None else AI_GATE_LOW_PROB

    if prob_5d >= threshold:
        return {"signal": signal, "advice": advice, "gate_triggered": False, "gate_threshold": threshold}

    horse_score = horse_result.get("horse_score", 0)

    # ── 階梯式 Override：由高分往低分比對，找到第一個同時滿足 (horse_score, ai_prob) 的層級 ──
    for tier_horse, tier_prob_min in AI_GATE_OVERRIDE_TIERS:
        if horse_score >= tier_horse and prob_5d >= tier_prob_min:
            gated_signal = "🚀 搶先布局(Horse Override)"
            gated_advice = (f"⚡ AI Decision Gate Override[{category}]｜AI機率{prob_5d:.1f}%(<動態門檻{threshold:.0f}%)，"
                             f"但HorseScore={horse_score}達覆寫層級(>={tier_horse}分 且 AI機率>={tier_prob_min}%)，允許搶先布局。原建議：{advice}")
            return {"signal": gated_signal, "advice": gated_advice, "gate_triggered": True, "gate_threshold": threshold}

    if signal in ["強力買入", "買入", "持平觀望"]:
        gated_signal = "🔍 觀察"
        gated_advice = (f"🚧 AI Decision Gate[{category}]｜AI機率僅{prob_5d:.1f}%(<動態門檻{threshold:.0f}%)，原訊號「{signal}」"
                         f"已鎖定為觀察等級，避免方向信心不足時過度積極。HorseScore={horse_score}(未達任何覆寫層級)。原建議：{advice}")
        return {"signal": gated_signal, "advice": gated_advice, "gate_triggered": True, "gate_threshold": threshold}

    # 原本已是風險控管訊號（減碼/賣出），Gate 不干預
    return {"signal": signal, "advice": advice, "gate_triggered": False, "gate_threshold": threshold}


def compute_final_decision(pred: dict, horse_result: dict, risk_regime: str) -> str:
    """
    V18.1：Final Decision Engine（決策統整層）
    把 AI Decision Gate 結果、HorseFinder、風險模式(Regime) 統一收斂成「唯一」的最終決策標籤，
    避免報表中出現「訊號寫買入、但AI機率很低、Horse也是NO」這種各欄位互相矛盾、看報表要自己拼湊的情況。
    優先序：風險模式(CRASH) > AI Decision Gate(Override/觀察鎖定) > 原始訊號(+Horse加成) > 一般訊號。
    """
    signal = pred.get("signal", "")
    gate_triggered = pred.get("gate_triggered", False)
    horse_flag = horse_result.get("horse_flag", False)

    if risk_regime == "CRASH":
        return "⚫ 全面出場（總體崩盤模式，不論個股訊號一律避險）"

    if gate_triggered:
        if "搶先布局" in signal:
            return f"🚀 積極卡位（AI信心不足但HorseScore={horse_result.get('horse_score',0)}極強，允許提前布局）"
        return f"🔍 觀察為主（AI信心不足，鎖定觀察，不追價）"

    if signal in ["強力買入", "買入"]:
        return (f"🔥 積極布局（AI與HorseFinder雙重確認）" if horse_flag else "🟢 正常買進（AI確認，Horse未達標）")
    if signal == "持平觀望":
        return "🟡 持平觀望（維持現狀，不加碼不減碼）"
    if signal in ["減碼", "賣出"]:
        return f"🔴 {signal}（風控訊號，優先執行）"
    return signal or "N/A"


def horse_breakdown_to_str(breakdown: dict) -> str:
    label_map = {
        "bb_squeeze": "BB擠壓", "bb_width_low10pct": "通道窄<10%分位",
        "volume_breakout": "股數爆量>2倍", "value_surge_ratio": "成交值爆量>2.5倍(資金流)",
        "new_high_20d": "20日新高",
        "macd_golden_cross": "MACD金叉", "adx_strong": "ADX>25",
        "institution_streak": "法人連買",
    }
    hits = [label_map[k] for k, v in breakdown.items() if v > 0]
    return "、".join(hits) if hits else "無顯著訊號"


# ==========================================
# 💰 V18：Kelly Position Sizing（半凱利，與 ATR / 風險倉位聯合取最小值）
# ==========================================
def kelly_fallback_by_ai_prob(prob_5d: float) -> dict:
    """
    V18.1：Kelly 樣本不足時的備用倉位，不再顯示生硬的 N/A，
    改依 AI 上漲機率分級給出保守的替代建議（明確標註為 Fallback，不與真實 Kelly 混淆）。
    """
    for threshold, frac in KELLY_FALLBACK_TIERS:
        if prob_5d >= threshold:
            return {"kelly_fraction": frac, "kelly_str": f"{frac * 100:.0f}%(Fallback)", "is_fallback": True}
    return {"kelly_fraction": KELLY_FALLBACK_TIERS[-1][1], "kelly_str": f"{KELLY_FALLBACK_TIERS[-1][1] * 100:.0f}%(Fallback)", "is_fallback": True}


def compute_kelly_position(records: list, prob_5d: float = 50.0) -> dict:
    """
    以個股歷史真實 Walk-Forward 交易紀錄估算 Kelly 最適倉位比例。
    f* = 勝率 - (1-勝率)/賠率 (賠率=平均獲利/平均虧損)
    採半凱利 (Half-Kelly) 降低過度槓桿風險，並設定硬上限。
    V18.1：樣本不足時不再回傳 N/A，改用 AI 上漲機率分級的 Fallback Kelly（並標註 is_fallback=True 以利報表區分）。
    """
    if not records or len(records) < KELLY_MIN_RECORDS:
        fb = kelly_fallback_by_ai_prob(prob_5d)
        fb["reason"] = "樣本不足"
        return fb

    rets = [r["strat_return"] for r in records]
    wins = [r for r in rets if r > 0]
    losses = [r for r in rets if r <= 0]
    win_rate = len(wins) / len(rets)
    avg_win = np.mean(wins) if wins else 0.0
    avg_loss = abs(np.mean(losses)) if losses else 0.0

    if avg_win <= 0 or avg_loss <= 0:
        fb = kelly_fallback_by_ai_prob(prob_5d)
        fb["reason"] = "無足夠正負樣本"
        return fb

    win_loss_ratio = avg_win / avg_loss
    kelly_f = win_rate - (1 - win_rate) / win_loss_ratio
    half_kelly = max(0.0, kelly_f * KELLY_HALF_FACTOR)
    kelly_capped = round(min(half_kelly, KELLY_MAX_FRACTION), 4)

    return {"kelly_fraction": kelly_capped, "kelly_str": f"{kelly_capped * 100:.1f}%", "is_fallback": False,
            "win_rate": round(win_rate * 100, 1), "win_loss_ratio": round(win_loss_ratio, 2)}



# ==========================================
# ⭐ V18：AI 信心等級（★ 星等）
# ==========================================
def compute_confidence_stars(pred: dict, val_data: dict, horse_result: dict, df: pd.DataFrame) -> dict:
    """
    綜合：AI上漲機率、多週期動能一致性、歷史回測勝率、黑馬訊號、BB Confidence Bonus
    產生 1~5 顆星的信心等級，方便看報表時一眼判斷。
    """
    score = 0
    prob_5d = pred.get("prob_5d", 50.0)
    if prob_5d >= 75: score += 2
    elif prob_5d >= 60: score += 1

    if pred.get("consistency_score", 0) >= 100:  # 4/4 週期同向
        score += 1

    hit_rate_str = val_data.get("hit_rate", "N/A")
    hit_rate_pct = 0.0
    try:
        if "(" in hit_rate_str:
            hit_rate_pct = float(hit_rate_str.split("(")[-1].replace("%)", ""))
    except Exception:
        pass
    if hit_rate_pct >= 60: score += 1

    if horse_result.get("horse_flag"): score += 1

    bb_bonus = 0
    if "BB_confidence_bonus" in df.columns and len(df) > 0:
        try: bb_bonus = int(df["BB_confidence_bonus"].iloc[-1])
        except Exception: bb_bonus = 0
    if bb_bonus > 0: score += 1

    stars_n = max(1, min(5, score if score > 0 else 1))
    return {"stars": "★" * stars_n + "☆" * (5 - stars_n), "stars_n": stars_n}


def predict_with_top_features(df: pd.DataFrame, top_features: list, best_params: dict, us_risk: dict = None,
                               stock_category: str = "default") -> dict:
    use_features = top_features[:TOP_N_FEATURES]
    
    X_last = df[use_features].iloc[[-1]].astype(np.float32).values
    current_price = float(df["close"].iloc[-1])
    latest_date = df.index[-1].strftime('%Y-%m-%d')
    
    atr_14 = current_price * 0.02
    if "ATR_14" in df.columns:
        valid_atr = df["ATR_14"].replace([np.inf, -np.inf], np.nan).dropna()
        if not valid_atr.empty:
            atr_14 = float(valid_atr.iloc[-1])
            if pd.isna(atr_14) or atr_14 <= 0: atr_14 = current_price * 0.02

    latest_adx = 20.0
    if "ADX" in df.columns:
        valid_adx = df["ADX"].replace([np.inf, -np.inf], np.nan).dropna()
        if not valid_adx.empty: latest_adx = float(valid_adx.iloc[-1])

    atr_pct = atr_14 / current_price
    risk_per_trade = 0.01
    position_size = (risk_per_trade / atr_pct) if atr_pct > 0 else 0
    position_size = min(position_size, 0.25)
    position_size_str = f"{position_size * 100:.1f}%"

    # ★ 專家建議：硬停損為進場價之 -6% 
    stop_loss = current_price * 0.94
    tp_cons = current_price + 1.8 * atr_14
    tp_agg = current_price + 4.0 * atr_14
    
    latest_regime = df["macro_regime"].iloc[-1] if "macro_regime" in df.columns else "低波動盤整"
    regime_label  = get_regime_label(latest_regime)
    # V18.2b：MARKET_MODE（BULL/SIDEWAYS/BEAR/PANIC），VIX 取自 us_risk（若尚未取得風險資料則以None處理，退回純macro判斷）
    _vix_for_mode = us_risk.get("vix_close") if (us_risk and us_risk.get("fetch_ok")) else None
    market_mode = classify_market_mode(latest_regime, _vix_for_mode)

    # 注入正規化參數防過度擬合
    xgb_params = {
        'n_estimators': 300,
        'max_depth': best_params.get('max_depth', 3),
        'learning_rate': best_params.get('learning_rate', 0.03),
        'min_child_weight': 8,
        'gamma': 0.5,
        'subsample': best_params.get('subsample', 0.7),
        'colsample_bytree': 0.7,
        'reg_alpha': best_params.get('reg_alpha', 0.5),
        'reg_lambda': best_params.get('reg_lambda', 2.0),
        'random_state': 42,
        'n_jobs': 4
    }
    
    lgb_params = {
        'n_estimators': 500, 
        'max_depth': best_params.get('max_depth', 4), 
        'learning_rate': best_params.get('learning_rate', 0.03), 
        'subsample': best_params.get('subsample', 0.8), 
        'colsample_bytree': 0.8, 
        'reg_alpha': best_params.get('reg_alpha', 0.5),
        'reg_lambda': best_params.get('reg_lambda', 2.0),
        'random_state': 42, 
        'n_jobs': 1, 
        'verbose': -1
    }

    preds, probs = {}, {}
    
    for days in HORIZONS:
        target_col = f"target_{days}d"
        target_up_col = f"target_up_{days}d"
        
        # 實時推論使用完整歷史資料訓練模型，不再不當截斷 recent 15% 資料
        df_clean = df[use_features + [target_col]].replace([np.inf, -np.inf], np.nan).dropna()
        if len(df_clean) < 60:
            preds[f"pred_{days}d"] = 0.0
        else:
            X_h = df_clean[use_features].astype(np.float32).values
            y_h = df_clean[target_col].astype(np.float32).values

            X_tr = X_h  # 使用 100% 數據訓練
            y_tr_clipped = np.clip(y_h, TARGET_CLIP_MIN, TARGET_CLIP_MAX)

            model_xgb = xgb.XGBRegressor(**xgb_params)
            sample_weight = np.linspace(0.3, 1.0, len(X_tr))
            model_xgb.fit(X_tr, y_tr_clipped, sample_weight=sample_weight, verbose=False)
            
            model_lgb = lgb.LGBMRegressor(**lgb_params)
            model_lgb.fit(X_tr, y_tr_clipped)

            preds[f"pred_{days}d"] = float((model_xgb.predict(X_last)[0] + model_lgb.predict(X_last)[0]) / 2)

            # ================= V18：SHAP 局部解釋（僅對 5 日核心預測計算，節省運算量） =================
            if days == FUTURE_DAYS and HAS_SHAP:
                try:
                    explainer = shap.TreeExplainer(model_xgb)
                    shap_vals_last = explainer.shap_values(X_last)
                    shap_row = np.array(shap_vals_last).reshape(-1)
                    order = np.argsort(-np.abs(shap_row))[:3]
                    preds["shap_top_features"] = [(use_features[i], round(float(shap_row[i]), 4)) for i in order]
                except Exception:
                    preds["shap_top_features"] = []

        # 上漲機率模型亦使用 100% 數據訓練
        df_cls = df[use_features + [target_up_col]].replace([np.inf, -np.inf], np.nan).dropna()
        up_prob_val = 50.0
        if len(df_cls) >= 60:
            try:
                X_c = df_cls[use_features].astype(np.float32).values
                y_c = df_cls[target_up_col].values
                X_c_tr, y_c_tr = X_c, y_c
                
                if len(np.unique(y_c_tr)) >= 2:
                    scale_pos_weight = max(1, (len(y_c_tr) - sum(y_c_tr)) / max(sum(y_c_tr), 1))
                    cls_xgb = xgb.XGBClassifier(**xgb_params, scale_pos_weight=scale_pos_weight, eval_metric="logloss")
                    cls_lgb = lgb.LGBMClassifier(**lgb_params, class_weight='balanced')
                    
                    cls_sample_weight = np.linspace(0.3, 1.0, len(X_c_tr))
                    cls_xgb.fit(X_c_tr, y_c_tr, sample_weight=cls_sample_weight, verbose=False)
                    cls_lgb.fit(X_c_tr, y_c_tr)
                    
                    up_prob_val = round(float((cls_xgb.predict_proba(X_last)[0][1] + cls_lgb.predict_proba(X_last)[0][1]) / 2) * 100, 1)
            except: pass
        probs[f"prob_{days}d"] = up_prob_val

    pred_1d, pred_3d, pred_5d, pred_20d = preds.get("pred_1d", 0.0), preds.get("pred_3d", 0.0), preds.get("pred_5d", 0.0), preds.get("pred_20d", 0.0)
    prob_1d, prob_3d, prob_5d, prob_20d = probs.get("prob_1d", 50.0), probs.get("prob_3d", 50.0), probs.get("prob_5d", 50.0), probs.get("prob_20d", 50.0)
    
    signs = [1 if p > 0 else -1 for p in [pred_1d, pred_3d, pred_5d, pred_20d]]
    pos_count = sum(1 for s in signs if s == 1)
    trend_stars = f"{pos_count}/4 {'★' * pos_count + '☆' * (4 - pos_count)}"
    consistency_score = pos_count * 25
    trend_score = compute_trend_score(pos_count, latest_adx)   # V18.2：供動態 AI Gate 門檻使用
    
    composite_score = (0.4 * pred_5d) + (0.3 * pred_20d) + (0.2 * prob_5d) + (0.1 * consistency_score)
    expected_profit = max(pred_5d, 0.1)
    expected_loss = max(((current_price - stop_loss) / current_price) * 100, 0.1)
    composite_score += (expected_profit / expected_loss) * 3

    pred_price = current_price * (1 + pred_5d / 100)

    # ★ 專家建議：10EMA 作為波段移動停利的防守價
    trailing_stop = round(float(df['EMA_10'].iloc[-1]), 2) if 'EMA_10' in df.columns else current_price

    latest_rsi = float(df["RSI_14"].iloc[-1]) if "RSI_14" in df.columns and pd.notna(df["RSI_14"].iloc[-1]) else 50.0

    base_result = {
        "pred_1d": round(pred_1d, 2), "pred_3d": round(pred_3d, 2), "pred_return": round(pred_5d, 2), "pred_20d": round(pred_20d, 2),
        "prob_1d": round(prob_1d, 1), "prob_3d": round(prob_3d, 1), "prob_5d": round(prob_5d, 1), "prob_20d": round(prob_20d, 1),
        "trend_stars": trend_stars, "consistency_score": consistency_score, "composite_score": round(composite_score, 2),
        "trend_score": trend_score, "stock_category": stock_category,   # V18.2
        "market_mode": market_mode, "latest_rsi": round(latest_rsi, 1),   # V18.2b
        "up_prob": prob_5d, "current_price": round(current_price, 2), "latest_date": latest_date,
        "stop_loss": round(stop_loss, 2), "tp_cons": round(tp_cons, 2), "tp_agg": round(tp_agg, 2), "trailing_stop": trailing_stop,
        "pred_price": round(pred_price, 2), "regime": regime_label, "position_size": position_size_str,
        "latest_adx": latest_adx,
        "atr_position_raw": position_size,   # V18：保留 ATR 原始倉位比例（浮點數），供 Kelly 聯合決策使用
        "shap_top_features": preds.get("shap_top_features", []),  # V18：5日預測之 SHAP 局部解釋（前3大貢獻因子）
    }

    # ========= V15/V18.2：預埋全球風險資訊到 base_result（V18.2 改用分業風險模型） =========
    if us_risk and us_risk.get("fetch_ok"):
        rs, rr, rd = compute_sector_risk_score(us_risk, stock_category)
        base_result["global_risk"] = {
            "score": rs, "regime": rr, "detail": rd, "raw": us_risk
        }

    # ★ 爆量黑馬波段起漲邏輯
    vol_break_today = df["Volume_Break"].iloc[-1]
    vol_break_yest  = df["Volume_Break"].iloc[-2] if len(df) > 1 else 0
    c_today, o_today = df["close"].iloc[-1], df["open"].iloc[-1]
    c_yest = df["close"].iloc[-2] if len(df) > 1 else c_today

    is_dark_horse = (vol_break_today == 1 or vol_break_yest == 1) and (c_today >= o_today or c_today >= c_yest) and pred_5d > 0.5

    # 1. 檢驗盤整區
    if abs(pred_5d) < 1.2 and prob_5d < 58 and latest_adx < 18:
        if is_dark_horse:
            base_result.update({"signal": "🔥 黑馬起漲", "advice": f"🔥 爆量黑馬｜剛脫離量縮盤整帶量突破，AI 確認趨勢動能轉強！預測 {pred_5d:.2f}%｜勝率 {prob_5d:.1f}%"})
            return base_result
        else:
            base_result.update({"signal": "觀望", "advice": "⚪ No Trade Zone｜盤整盤 + 低趨勢強度，避免交易"})
            return base_result

    # 2. 趨勢與買賣決策設定
    if pred_5d >= 3.0 and pos_count >= 3:  signal = "強力買入"
    elif pred_5d >= 1.0 or pos_count >= 3: signal = "買入"
    elif pred_5d >= -1.0:                  signal = "持平觀望"
    elif pred_5d >= -3.0:                  signal = "減碼"
    else:                                  signal = "賣出"

    # 3. 結合大盤多空過濾器
    if latest_regime in ["崩盤", "主跌段"] and signal in ["強力買入", "買入", "持平觀望", "🔥 黑馬起漲"]:
        signal, advice = "減碼", f"🔴 Regime 崩盤警告｜大盤破季線，強制降至【減碼】"
    elif latest_regime in ["空頭反彈"] and signal in ["強力買入", "買入", "🔥 黑馬起漲"]:
        signal, advice = "持平觀望", f"🟠 Regime 空頭環境｜大盤反彈結構，強制降至【持平觀望】"
    else:
        prob_str = f"｜綜合勝率 {prob_5d:.1f}%"
        if signal == "強力買入": advice = f"🟢 強力買入｜多週期同步轉強，預測上漲 {pred_5d:.2f}%{prob_str}"
        elif signal == "買入":   advice = f"🔵 買入觀察｜趨勢偏多，預測上漲 {pred_5d:.2f}%{prob_str}"
        elif signal == "持平觀望": advice = f"🟡 持平觀望｜長短週期震盪，預測漲跌 {pred_5d:.2f}%{prob_str}"
        elif signal == "減碼":   advice = f"🟠 建議減碼｜趨勢轉弱，預測下跌 {pred_5d:.2f}%{prob_str}"
        else:                    advice = f"🔴 建議賣出｜多週期同步走弱，預測下跌 {pred_5d:.2f}%{prob_str}"

    # 4. 起漲突破高量催化直接晉升黑馬
    if is_dark_horse and signal in ["強力買入", "買入"]:
        signal = "🔥 黑馬起漲"
        advice = f"🔥 爆量黑馬｜剛脫離量縮盤整帶量突破，AI 確認趨勢動能轉強！預測 {pred_5d:.2f}%{prob_str}"

    # ========= V15 核心：全球風險強制覆寫 =========
    # 此層在所有 AI 技術信號之後執行，不可被規避
    # risk 字典由外部傳入（predict_with_top_features 函式簽名已更新）
    if "global_risk" in base_result:
        risk_info = base_result.pop("global_risk")
        risk_score_val  = risk_info.get("score", 0)
        risk_regime_val = risk_info.get("regime", "NORMAL")
        risk_detail_val = risk_info.get("detail", "")
        risk_raw        = risk_info.get("raw", {})

        signal, advice, pred_5d_adj, composite_adj, max_pos_override, max_pos_raw = apply_global_risk_override(
            signal, advice, pred_5d, prob_5d,
            base_result["composite_score"],
            risk_score_val, risk_regime_val, risk_detail_val
        )
        base_result["composite_score"] = round(composite_adj, 2)
        base_result["pred_return"]     = round(pred_5d_adj, 2)   # 調整後的預測報酬
        base_result["risk_score"]      = risk_score_val
        base_result["risk_regime"]     = risk_regime_val
        base_result["risk_detail"]     = risk_detail_val
        base_result["vix_close"]       = risk_raw.get("vix_close", 18.0)
        base_result["nasdaq_ret1"]     = risk_raw.get("nasdaq_ret1", 0.0)
        base_result["sox_ret1"]        = risk_raw.get("sox_ret1", 0.0)
        # V16：以風險調整後的倉位覆寫原始 ATR 倉位
        base_result["position_size"]   = max_pos_override
        base_result["regime_position_raw"] = max_pos_raw   # V18：保留數值化的風險倉位上限，供 Kelly 聯合決策使用

    base_result.update({"signal": signal, "advice": advice})
    return base_result

def compute_trade_stats(records: list) -> dict:
    """
    V18.4：專家建議 —— 勝率不夠，Profit Factor 與 Expectancy 才是專業交易更重要的指標。
      Gross Profit  = 所有獲利交易的報酬總和
      Gross Loss    = 所有虧損交易的報酬總和(取絕對值)
      Profit Factor = Gross Profit / Gross Loss（>1 代表整體是賺錢的，越高越好；>=2 通常視為穩健策略）
      Expectancy(E) = 勝率 × 平均獲利 - 敗率 × 平均虧損（=每筆交易的平均期望報酬%，與『所有交易報酬取平均』數學上等價）
    輸入 records 為 historical_validation() 產生的交易紀錄列表（每筆含 strat_return 欄位，單位:%）。
    若沒有任何虧損交易，Profit Factor 視為 None（代表無限大，避免除以0），前端顯示為「∞(無虧損)」。
    """
    if not records:
        return {"gross_profit": 0.0, "gross_loss": 0.0, "profit_factor": None, "expectancy": 0.0,
                "avg_win": 0.0, "avg_loss": 0.0, "win_rate_pf": 0.0, "num_wins": 0, "num_losses": 0}

    rets = [r["strat_return"] for r in records]
    wins = [r for r in rets if r > 0]
    losses = [r for r in rets if r <= 0]

    gross_profit = round(sum(wins), 2)
    gross_loss = round(abs(sum(losses)), 2)
    profit_factor = round(gross_profit / gross_loss, 2) if gross_loss > 0 else None

    avg_win = round(float(np.mean(wins)), 2) if wins else 0.0
    avg_loss = round(float(abs(np.mean(losses))), 2) if losses else 0.0
    win_rate_pf = len(wins) / len(rets) if rets else 0.0
    lose_rate_pf = 1 - win_rate_pf

    # Expectancy：兩種算法數學上等價，直接用「全部交易報酬平均」最穩健（樣本數少時，WinRate×AvgWin公式可能因四捨五入產生些微差異）
    expectancy = round(float(np.mean(rets)), 2) if rets else 0.0

    return {"gross_profit": gross_profit, "gross_loss": gross_loss, "profit_factor": profit_factor,
            "expectancy": expectancy, "avg_win": avg_win, "avg_loss": avg_loss,
            "win_rate_pf": round(win_rate_pf * 100, 1), "num_wins": len(wins), "num_losses": len(losses)}


def format_profit_factor(pf) -> str:
    """把 profit_factor(可能為 None=無虧損交易) 轉成報表友善字串"""
    if pf is None:
        return "∞(無虧損交易)"
    return f"{pf:.2f}"


def historical_validation(df: pd.DataFrame, top_features: list, best_params: dict) -> dict:
    xgb_params = {
        'n_estimators': 150, 'max_depth': best_params.get('max_depth', 3), 'learning_rate': best_params.get('learning_rate', 0.05),
        'reg_alpha': best_params.get('reg_alpha', 0.5), 'reg_lambda': best_params.get('reg_lambda', 2.0),
        'random_state': 42, 'n_jobs': 4
    }
    lgb_params = {
        'n_estimators': 200, 'max_depth': best_params.get('max_depth', 4), 'learning_rate': best_params.get('learning_rate', 0.05),
        'reg_alpha': best_params.get('reg_alpha', 0.5), 'reg_lambda': best_params.get('reg_lambda', 2.0),
        'random_state': 42, 'verbose': -1, 'n_jobs': 1
    }

    use_features = top_features[:TOP_N_FEATURES]
    needed_cols  = list(dict.fromkeys(use_features + ["target_5d", "target_up_5d", "close", "ADX", "macro_regime", "Volume_Break", "open", "low", "EMA_10", "Low_5", "volume", "MACD_hist"]))
    avail_cols   = [c for c in needed_cols if c in df.columns]
    df_clean = df[avail_cols].replace([np.inf, -np.inf], np.nan).dropna()
    
    records, strat_returns_list = [], []
    if len(df_clean) < 150: return {"records": records, "sharpe": 0.0, "mdd": 0.0, "hit_rate": "N/A", **compute_trade_stats([])}
    
    X, y_reg, y_cls = df_clean[use_features], df_clean["target_5d"], df_clean["target_up_5d"]

    # 1. 預先計算歷史評估點
    test_len = 12 * FUTURE_DAYS 
    start_eval = max(100, len(df_clean) - test_len)
    eval_points = list(range(start_eval, len(df_clean), FUTURE_DAYS))

    pred_map = {}
    for end_idx in eval_points:
        X_tr, y_tr_reg, y_tr_cls = X.iloc[:end_idx-FUTURE_DAYS].values, y_reg.iloc[:end_idx-FUTURE_DAYS].values, y_cls.iloc[:end_idx-FUTURE_DAYS].values
        X_test = X.iloc[[end_idx]].values
        
        m_x, m_l = xgb.XGBRegressor(**xgb_params), lgb.LGBMRegressor(**lgb_params)
        m_x.fit(X_tr, y_tr_reg)
        m_l.fit(X_tr, y_tr_reg)
        pred_ret = float((m_x.predict(X_test)[0] + m_l.predict(X_test)[0]) / 2)

        up_prob_val = 50.0
        if len(np.unique(y_tr_cls)) >= 2:
            c_x, c_l = xgb.XGBClassifier(**xgb_params, eval_metric="logloss"), lgb.LGBMClassifier(**lgb_params, class_weight='balanced')
            c_x.fit(X_tr, y_tr_cls)
            c_l.fit(X_tr, y_tr_cls)
            up_prob_val = float((c_x.predict_proba(X_test)[0][1] + c_l.predict_proba(X_test)[0][1]) / 2) * 100
            
        pred_map[end_idx] = {"pred_ret": pred_ret, "up_prob": up_prob_val}

    # 2. 逐日波段回測狀態機 (讓獲利奔跑，直到趨勢跌破)
    in_pos = False
    entry_price, entry_idx = 0.0, 0
    current_pred_ret, current_up_prob = 0.0, 50.0

    for i in range(start_eval, len(df_clean)):
        if i in pred_map:
            current_pred_ret = pred_map[i]["pred_ret"]
            current_up_prob = pred_map[i]["up_prob"]

        if not in_pos:
            vol_break_today = df_clean["Volume_Break"].iloc[i]
            vol_break_yest  = df_clean["Volume_Break"].iloc[i - 1] if i > 0 else 0
            c_today, o_today = df_clean["close"].iloc[i], df_clean["open"].iloc[i]
            c_yest = df_clean["close"].iloc[i - 2] if i > 1 else c_today # 這裡修正為安全索引範圍
            
            is_dark_horse = (vol_break_today == 1 or vol_break_yest == 1) and (c_today >= o_today or c_today >= c_yest) and current_pred_ret > 0.5
            
            if (current_pred_ret > 1.0 and current_up_prob > 55) or is_dark_horse:
                in_pos = True
                entry_idx = i
                entry_price = c_today
        else:
            c = df_clean['close'].iloc[i]
            o = df_clean['open'].iloc[i]
            vol = df_clean['volume'].iloc[i]
            vol_ma20 = df_clean['volume'].rolling(20).mean().iloc[i] if i >= 20 else vol
            
            ema10 = df_clean['EMA_10'].iloc[i]
            macd_h = df_clean['MACD_hist'].iloc[i]
            adx_today = df_clean['ADX'].iloc[i]
            adx_yest = df_clean['ADX'].iloc[i-1]
            low_5 = df_clean['Low_5'].iloc[i-1]
            macro = df_clean['macro_regime'].iloc[i] if 'macro_regime' in df_clean.columns else ""

            cond_sl = (c < entry_price * 0.94)
            cond_trend_break = (c < ema10) and (macd_h < 0 or (adx_today < adx_yest and adx_today > 25))
            cond_low_break = (c < low_5)
            cond_panic = (vol > vol_ma20 * 2) and (c < o * 0.96)
            cond_macro = (macro in ['崩盤', '主跌段'])
            
            exit_triggered = cond_sl or cond_trend_break or cond_low_break or cond_panic or cond_macro
            is_last_day = (i == len(df_clean) - 1)

            if exit_triggered or is_last_day:
                exit_price = c
                raw_ret   = (exit_price / entry_price - 1) * 100
                # V16：backtest 內套用動態倉位（以 backtest 期間 macro 環境估算）
                # 此處簡化：使用當日 macro_regime 對應倉位縮放報酬
                macro_now = df_clean["macro_regime"].iloc[i] if "macro_regime" in df_clean.columns else "低波動盤整"
                # 保守估算：崩盤期=50%倉，主跌段=70%倉，其他=100%倉
                position_scale = 0.50 if macro_now == "崩盤" else (0.70 if macro_now == "主跌段" else 1.0)
                trade_ret = raw_ret * position_scale - COST_PCT
                hold_days = i - entry_idx
                
                if cond_sl: reason = "強制停損"
                elif cond_panic: reason = "爆量長黑"
                elif cond_low_break: reason = "跌破前低"
                elif cond_trend_break: reason = "破線(10EMA)"
                elif cond_macro: reason = "大盤轉空"
                else: reason = "期末平倉"

                strat_returns_list.append(trade_ret / 100)
                records.append({
                    "period": f"Rolling {df_clean.index[end_idx].strftime('%m-%d')}" if 'end_idx' in locals() else "Rolling",
                    "date": df_clean.index[i].strftime("%Y-%m-%d"),
                    "entry_date": df_clean.index[entry_idx].strftime('%Y-%m-%d'),
                    "exit_date": df_clean.index[i].strftime('%Y-%m-%d'),
                    "hold_days": hold_days,
                    "pred_return": round(current_pred_ret, 2),
                    "actual_return": round((c / entry_price - 1) * 100, 2),
                    "strat_return": round(trade_ret, 2),
                    "is_hit": "✅ 命中" if trade_ret > 0 else "❌ 失誤",
                    "is_hit_bool": trade_ret > 0,
                    "reason": reason
                })
                in_pos = False

    sharpe, mdd, hit_rate = 0.0, 0.0, "N/A"
    return_std = 0.0
    if strat_returns_list:
        returns = np.array(strat_returns_list)
        return_std = round(float(returns.std() * 100), 3)
        if returns.std() > 0:
            avg_hold = max(1, np.mean([r["hold_days"] for r in records]))
            sharpe = round((returns.mean() / (returns.std() + 1e-9)) * np.sqrt(252 / avg_hold), 2)
        cum = np.cumsum(returns)
        peak = np.maximum.accumulate(cum)
        mdd = round((cum - peak).min() * 100, 2)
        hits = sum(1 for r in records if r["is_hit_bool"])
        hit_rate = f"{hits}/{len(records)} ({hits/len(records)*100:.0f}%)"

    # V18：正式標註為 Walk-Forward Validation（Anchored Expanding Window），並回傳折數與報酬標準差供報表展示
    trade_stats = compute_trade_stats(records)   # V18.4：Profit Factor / Expectancy
    return {"records": records, "sharpe": sharpe, "mdd": mdd, "hit_rate": hit_rate,
            "n_folds": len(eval_points), "return_std": return_std, **trade_stats}

def setup_excel_styles(wb):
    if "header_style" not in wb.named_styles:
        header_style = NamedStyle(name="header_style")
        header_style.font = Font(bold=True, color=COLORS["header_fg"], size=10, name="Arial")
        header_style.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
        header_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        wb.add_named_style(header_style)
        
    if "basic_border" not in wb.named_styles:
        basic_border = NamedStyle(name="basic_border")
        basic_border.border = Border(left=Side(style="thin", color=COLORS["border"]), right=Side(style="thin", color=COLORS["border"]),
                                     top=Side(style="thin", color=COLORS["border"]), bottom=Side(style="thin", color=COLORS["border"]))
        basic_border.alignment = Alignment(horizontal="center", vertical="center")
        wb.add_named_style(basic_border)

def set_col_widths(ws, widths: dict):
    for col_letter, w in widths.items(): ws.column_dimensions[col_letter].width = w

def clean_sheet_name(stock_id: str, company: str) -> str:
    name = f"{stock_id} {company}"
    for ch in INVALID_SHEET_CHARS: name = name.replace(ch, "")
    return name[:31]

def write_stock_sheet(wb, sheet_name, stock_id, company, all_features, effective, df_price, pred_result, val_data,
                       horse_result=None, kelly_info=None, final_position_str=None, confidence=None, final_decision=None):
    horse_result = horse_result or {"horse_score": 0, "horse_flag": False, "breakdown": {}, "horse_stars": "★☆☆☆☆", "horse_tier": "極弱"}
    kelly_info = kelly_info or {"kelly_str": "N/A"}
    confidence = confidence or {"stars": "☆☆☆☆☆", "stars_n": 0}
    final_decision = final_decision or pred_result.get("signal", "N/A")
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:K1")
    ws["A1"].value = f"📊 {stock_id} {company} — 法人級多週期預測與波段防守報告"
    ws["A1"].font  = Font(bold=True, size=14, color="FFFFFF", name="Arial")
    ws["A1"].fill  = PatternFill("solid", fgColor=COLORS["header_bg"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    for addr, val, bold in [
        ("A2", "分析日期", True), ("B2", datetime.now().strftime("%Y-%m-%d"), False),
        ("D2", "資料筆數", True), ("E2", len(df_price), False),
        ("G2", "有效指標數", True), ("H2", len(effective), False),
        ("I2", f"{df_price.index[0].strftime('%Y-%m-%d')} ~ {df_price.index[-1].strftime('%Y-%m-%d')}", False)
    ]:
        ws[addr] = val
        ws[addr].font = Font(bold=bold, name="Arial", size=10, color="000000" if bold else "1F497D")
    ws.merge_cells("I2:K2")

    # 專家波段防守與決策關鍵價位 (完全取代死板的5日目標價)
    ws.merge_cells("A4:D4")
    ws["A4"].value = "▌ 波段防守與出場決策關鍵參考價"
    ws["A4"].font  = Font(bold=True, size=11, color="FFFFFF", name="Arial")
    ws["A4"].fill  = PatternFill("solid", fgColor="4A235A")

    headers_def = ["指標 / 價位類型", "關鍵參考價", "目前差距(%)", "出場觸發條件"]
    for ci, h in enumerate(headers_def, 1):
        cell = ws.cell(row=5, column=ci, value=h)
        cell.style = "header_style"
        cell.fill = PatternFill("solid", fgColor="6C3483")

    cur_px = pred_result["current_price"]
    ema10_val = round(float(df_price['EMA_10'].iloc[-1]), 2) if 'EMA_10' in df_price.columns else cur_px
    low5_val = round(float(df_price['Low_5'].iloc[-1]), 2) if 'Low_5' in df_price.columns else cur_px
    stop_loss_val = pred_result.get("stop_loss", cur_px * 0.94)
    vol_ma20 = df_price['volume'].rolling(20).mean().iloc[-1] if len(df_price) >= 20 else df_price['volume'].iloc[-1]
    vol_break_val = round(vol_ma20 * 2, 0)

    pct_to_ema10 = round((cur_px / ema10_val - 1) * 100, 2) if ema10_val > 0 else 0.0
    pct_to_low5 = round((cur_px / low5_val - 1) * 100, 2) if low5_val > 0 else 0.0

    def_rows = [
        ("10EMA 趨勢線", ema10_val, f"{pct_to_ema10:+.2f}%", "收盤跌破 且 MACD翻黑/ADX下降"),
        ("近5日低點(支撐)", low5_val, f"{pct_to_low5:+.2f}%", "收盤跌破 (波段結構轉弱)"),
        ("硬停損點位 (-6%)", stop_loss_val, "-6.00%", "盤中或收盤跌破 (絕對停損出場)"),
        ("爆量警戒成交量", f"{vol_break_val:,.0f} 股", "N/A", "量能超過此警戒且收盤大跌K線")
    ]

    for ri, (title, val, gap, cond) in enumerate(def_rows, 6):
        ws.cell(row=ri, column=1, value=title).style = "basic_border"
        
        c_val = ws.cell(row=ri, column=2, value=val)
        c_val.style = "basic_border"
        c_val.font = Font(bold=True, name="Arial")
        
        c_gap = ws.cell(row=ri, column=3, value=gap)
        c_gap.style = "basic_border"
        if "-" in str(gap) and gap != "-6.00%":
            c_gap.font = Font(color="C0392B", bold=True, name="Arial")
        elif "+" in str(gap):
            c_gap.font = Font(color="27AE60", bold=True, name="Arial")
        
        ws.cell(row=ri, column=4, value=cond).style = "basic_border"

    sig_colors = {"強力買入": "1B5E20", "買入": "2E7D32", "觀望": "34495E", "持平觀望": "E65100", "減碼": "BF360C", "賣出": "B71C1C", "🔥 黑馬起漲": "D35400"}
    sig = pred_result.get("signal", "")

    # V15：在個股頁顯示全球風險狀態
    risk_regime_cell = pred_result.get("risk_regime", "NORMAL")
    risk_emoji_cell  = {"NORMAL": "🟢", "CAUTION": "🟡", "HIGH_RISK": "🔴", "CRASH": "⚫"}.get(risk_regime_cell, "⚪")
    vix_cell         = pred_result.get("vix_close", 18.0)
    nasdaq_cell      = pred_result.get("nasdaq_ret1", 0.0)
    sox_cell         = pred_result.get("sox_ret1", 0.0)
    risk_score_cell  = pred_result.get("risk_score", 0)
    risk_detail_cell = pred_result.get("risk_detail", "全球市場正常")

    ws.merge_cells("F5:K5")
    ws["F5"] = (f"📌 收盤日期：{pred_result.get('latest_date', '')} ｜ 目前收盤價：{cur_px} ｜ 市場環境：{pred_result['regime']} ｜ "
                f"{risk_emoji_cell} 全球風險:{risk_regime_cell} 量化分數:{risk_score_cell:.0f}/100 "
                f"| VIX:{vix_cell:.1f} NASDAQ:{nasdaq_cell:+.2f}% SOX:{sox_cell:+.2f}%")
    ws["F5"].font = Font(bold=True, size=10, color="1F497D", name="Arial")
    ws["F5"].fill = PatternFill("solid", fgColor="EBF5FB")
    ws["F5"].alignment = Alignment(vertical="center", indent=1)

    ws.merge_cells("F6:K6")
    ws["F6"] = f"多週期動能：{pred_result['trend_stars']} ｜ 交易信號：{sig} ｜ 資金配比建議：{pred_result.get('position_size', 'N/A')}"
    ws["F6"].font = Font(bold=True, size=10, color=sig_colors.get(sig, "000000"), name="Arial")
    ws["F6"].fill = PatternFill("solid", fgColor="EBF5FB")
    ws["F6"].alignment = Alignment(vertical="center", indent=1)

    ws.merge_cells("F7:K7")
    ws["F7"] = f"量化波段評分：{pred_result['composite_score']} 分 (5D/20D/勝率及動能加權)"
    ws["F7"].font = Font(bold=True, size=10, color="8E44AD", name="Arial")
    ws["F7"].fill = PatternFill("solid", fgColor="EBF5FB")
    ws["F7"].alignment = Alignment(vertical="center", indent=1)

    ws.merge_cells("F8:K8")
    ws["F8"] = f"🎯 動態防守：{pred_result.get('trailing_stop', '')} (10EMA) ｜ 保守目標：讓獲利奔跑 ｜ 積極目標：波段大趨勢"
    ws["F8"].font = Font(bold=True, size=10, color="C0392B", name="Arial")
    ws["F8"].fill = PatternFill("solid", fgColor="FADBD8")
    ws["F8"].alignment = Alignment(vertical="center", indent=1)

    ws.merge_cells("F9:K10")
    ws["F9"] = f"系統建議：{pred_result['advice']}"
    ws["F9"].font = Font(bold=True, size=10, name="Arial")
    if sig == "觀望":
        ws["F9"].fill = PatternFill("solid", fgColor="EAEDED")
    else:
        ws["F9"].fill = PatternFill("solid", fgColor="FDEDEC" if sig in ["賣出", "減碼"] else ("E9F7EF" if sig in ["買入", "強力買入"] else ("FFF2CC" if sig == "🔥 黑馬起漲" else "FEF9E7")))
    ws["F9"].alignment = Alignment(wrap_text=True, vertical="center", indent=1)

    for r in range(5, 11):
        for c in range(6, 12):
            ws.cell(row=r, column=c).border = Border(left=Side(style="thin", color="BDD7EE"), right=Side(style="thin", color="BDD7EE"),
                                                     top=Side(style="thin", color="BDD7EE"), bottom=Side(style="thin", color="BDD7EE"))

    # ================= V18.1：AI 信心等級面板（一眼看懂，Horse 改為星等分級，新增 Decision Gate 狀態） =================
    bb_squeeze_now = int(df_price["BB_squeeze"].iloc[-1]) if "BB_squeeze" in df_price.columns and len(df_price) > 0 else 0
    macro_now = df_price["macro_regime"].iloc[-1] if "macro_regime" in df_price.columns and len(df_price) > 0 else "N/A"
    macro_status_str = "GOOD" if macro_now in ["強趨勢多頭", "緩漲盤"] else ("BAD" if macro_now in ["崩盤", "主跌段", "空頭反彈"] else "NEUTRAL")
    risk_level_str = {"NORMAL": "LOW", "CAUTION": "MEDIUM", "HIGH_RISK": "HIGH", "CRASH": "EXTREME"}.get(risk_regime_cell, "LOW")
    horse_stars_disp = horse_result.get("horse_stars", "★☆☆☆☆")
    horse_tier_disp = horse_result.get("horse_tier", "極弱")

    ws.merge_cells("A11:D11")
    ws["A11"] = (f"⭐ AI 信心等級｜AI Score：{pred_result['composite_score']}｜Confidence：{confidence['stars']}｜"
                 f"Horse：{horse_stars_disp}({horse_result['horse_score']}分/{horse_tier_disp})｜"
                 f"Regime：{risk_regime_cell}｜Risk：{risk_level_str}｜BB Squeeze：{'YES' if bb_squeeze_now else 'NO'}｜Macro：{macro_status_str}｜"
                 f"類股：{pred_result.get('stock_category','default')}｜MARKET_MODE：{pred_result.get('market_mode','SIDEWAYS')}｜"
                 f"資金輪動：{pred_result.get('sector_tag','其他')}熱度{pred_result.get('sector_heat',50.0):.0f}分｜"
                 f"FinalScore：{pred_result.get('final_score_with_heat', pred_result['composite_score'])}")
    ws["A11"].font = Font(bold=True, size=10, color="7D3C98", name="Arial")
    ws["A11"].fill = PatternFill("solid", fgColor="F4ECF7")
    ws["A11"].alignment = Alignment(wrap_text=True, vertical="center", indent=1)

    gate_triggered = pred_result.get("gate_triggered", False)
    gate_threshold_disp = pred_result.get("gate_threshold", AI_GATE_LOW_PROB)
    if gate_triggered:
        gate_line = f"🚧 AI Decision Gate 已觸發：原訊號「{pred_result.get('original_signal', '')}」→ 現訊號「{pred_result.get('signal', '')}」（AI機率<動態門檻{gate_threshold_disp:.0f}%，HorseScore={horse_result['horse_score']}）"
        gate_color = "C0392B"
    else:
        gate_line = f"✅ AI Decision Gate 未觸發（AI機率>=動態門檻{gate_threshold_disp:.0f}% 或原訊號已屬風控類，無需修正）"
        gate_color = "1B5E20"
    if pred_result.get("mode_adjusted", False):
        gate_line += f" ｜🎯 MARKET_MODE={pred_result.get('market_mode','')}風格過濾已介入（詳見下方建議說明）"
    ws.merge_cells("A12:D12")
    ws["A12"] = gate_line
    ws["A12"].font = Font(bold=True, size=9, color=gate_color, name="Arial")
    ws["A12"].fill = PatternFill("solid", fgColor="FDEDEC" if gate_triggered else "EAFAF1")
    ws["A12"].alignment = Alignment(wrap_text=True, vertical="center", indent=1)

    kelly_str = kelly_info.get("kelly_str", "N/A")
    kelly_note = "（Fallback：依AI機率分級估算，非真實Kelly）" if kelly_info.get("is_fallback") else "（真實 Walk-Forward 樣本計算）"
    final_pos_display = final_position_str or pred_result.get("position_size", "N/A")
    shap_feats = pred_result.get("shap_top_features", [])
    shap_str = "、".join([f"{f}({v:+.3f})" for f, v in shap_feats]) if shap_feats else "N/A（未安裝 shap 或樣本不足）"

    ws.merge_cells("F11:K11")
    ws["F11"] = f"💰 Kelly建議倉位：{kelly_str}{kelly_note}｜綜合建議倉位(ATR/風險/Kelly取最小)：{final_pos_display}"
    ws["F11"].font = Font(bold=True, size=10, color="1B5E20", name="Arial")
    ws["F11"].fill = PatternFill("solid", fgColor="E8F8F5")
    ws["F11"].alignment = Alignment(vertical="center", indent=1)

    ws.merge_cells("A13:K13")
    ws["A13"] = f"🎯 Final Decision Engine（統整最終決策）：{final_decision}"
    ws["A13"].font = Font(bold=True, size=11, color="FFFFFF", name="Arial")
    ws["A13"].fill = PatternFill("solid", fgColor="34495E")
    ws["A13"].alignment = Alignment(wrap_text=True, vertical="center", indent=1)
    ws.row_dimensions[13].height = 20

    ws.merge_cells("F12:K12")
    ws["F12"] = f"🧠 SHAP 局部解釋(5日預測前3大貢獻因子)：{shap_str}"
    ws["F12"].font = Font(bold=True, size=10, color="2874A6", name="Arial")
    ws["F12"].fill = PatternFill("solid", fgColor="EBF5FB")
    ws["F12"].alignment = Alignment(wrap_text=True, vertical="center", indent=1)

    for r in range(11, 13):
        for c in range(1, 12):
            ws.cell(row=r, column=c).border = Border(left=Side(style="thin", color="BDD7EE"), right=Side(style="thin", color="BDD7EE"),
                                                     top=Side(style="thin", color="BDD7EE"), bottom=Side(style="thin", color="BDD7EE"))

    HR = 15
    ws.merge_cells(f"A{HR}:M{HR}")
    ws[f"A{HR}"].value = f"▌ 高級核心決策指標排名（已剔除共線性因子，V18 新增 SHAP 全域重要度）"
    ws[f"A{HR}"].font  = Font(bold=True, size=11, color="FFFFFF", name="Arial")
    ws[f"A{HR}"].fill  = PatternFill("solid", fgColor=COLORS["subheader"])

    for ci, h in enumerate(["排名", "指標名稱", "XGBoost重要性", "LightGBM重要性", "平均重要性", "SHAP重要度",
                              "Pearson相關", "Spearman相關", "絕對相關", "互信息MI",
                              "綜合評分", "是否有效", "決策核心"], 1):
        cell = ws.cell(row=HR + 1, column=ci, value=h)
        cell.style = "header_style"
        cell.fill = PatternFill("solid", fgColor=COLORS["subheader"])

    eff_set  = set(effective["feature"].tolist()) if effective is not None else set()
    top10_set = set(effective["feature"].head(10).tolist()) if effective is not None else set()

    for ri, row_data in all_features.iterrows():
        is_t10, is_e = row_data["feature"] in top10_set, row_data["feature"] in eff_set
        vals = [row_data["rank"], row_data["feature"], round(row_data["xgb_importance"], 4), round(row_data["lgb_importance"], 4),
                round(row_data["avg_importance"], 4), round(row_data.get("shap_importance", 0), 4),
                round(row_data["pearson"], 4), round(row_data["spearman"], 4), round(row_data["abs_pearson"], 4),
                round(row_data.get("mutual_info", 0), 4), round(row_data["final_score"], 4), "✅ 有效" if is_e else "❌ 篩除", "🌟 Top 10" if is_t10 else ""]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=HR + 2 + ri, column=ci, value=v)
            cell.style = "basic_border"
            cell.fill = PatternFill("solid", fgColor=("FFD700" if is_t10 else (COLORS["green_bg"] if is_e else ("FFFFFF" if ri % 2 == 0 else COLORS["alt_row"]))))
            cell.font = Font(name="Arial", size=10, bold=(ci in [1, 2] and is_t10))
            if ci == 2: cell.alignment = Alignment(horizontal="left", vertical="center")
            if ci in range(3, 12): cell.number_format = "0.0000"

    # 趨勢圖繪製
    try:
        plot_df = df_price.tail(60).copy()
        if len(plot_df) > 1:
            close_vals  = plot_df["close"].values.astype(float)
            valid_mask = ~np.isnan(close_vals) & ~np.isinf(close_vals)
            if not valid_mask.any():
                close_vals = np.full(len(close_vals), df_price["close"].dropna().iloc[-1] if not df_price["close"].dropna().empty else 0.0)
            else:
                s = pd.Series(close_vals).replace([np.inf, -np.inf], np.nan)
                close_vals = s.interpolate(method='linear').bfill().ffill().values

            x = np.arange(len(close_vals))
            coeffs = np.polyfit(x, close_vals, 1)
            trend_vals = np.polyval(coeffs, x)
            
            ws.cell(row=1, column=27, value="Date")
            ws.cell(row=1, column=28, value="收盤價 (TWD)")
            ws.cell(row=1, column=29, value="線性趨勢 (TWD)")
            
            for i, (date_idx, _) in enumerate(plot_df.iterrows()):
                r = i + 2
                ws.cell(row=r, column=27, value=date_idx.strftime("%Y-%m-%d"))
                ws.cell(row=r, column=28, value=close_vals[i])
                ws.cell(row=r, column=29, value=trend_vals[i])
                
            chart = LineChart()
            chart.title = f"{stock_id} {company} 近3月收盤價與趨勢"
            chart.style = 10
            chart.y_axis.title = '收盤價 (TWD)'
            chart.x_axis.title = '日期'
            chart.x_axis.tickLblSkip = 5  
            chart.width = 24
            chart.height = 14
            
            min_p, max_p = float(np.min(close_vals)), float(np.max(close_vals))
            margin = (max_p - min_p) * 0.15 if max_p > min_p else min_p * 0.05
            chart.y_axis.scaling.min = max(0.0, min_p - margin)
            chart.y_axis.scaling.max = max_p + margin
            chart.y_axis.number_format = '#,##0.0'

            last_row_data = len(plot_df) + 1
            dates_ref = Reference(ws, min_col=27, min_row=2, max_row=last_row_data)
            close_ref = Reference(ws, min_col=28, min_row=1, max_row=last_row_data)
            trend_ref = Reference(ws, min_col=29, min_row=1, max_row=last_row_data)

            chart.add_data(close_ref, titles_from_data=True)
            chart.add_data(trend_ref, titles_from_data=True)
            chart.set_categories(dates_ref)
            
            if len(chart.series) >= 2:
                chart.series[0].graphicalProperties.line.solidFill = "0070C0" 
                chart.series[0].smooth = True
                chart.series[1].graphicalProperties.line.solidFill = "FF4500" 
                chart.series[1].graphicalProperties.line.dashDot = "dash"

            ws.add_chart(chart, "N11")
    except Exception as e:
         print(f"    ⚠️ 趨勢圖繪製異常: {e}")

    val_records = val_data.get("records", [])
    if val_records:
        SC = HR + len(all_features) + 4
        ws.merge_cells(f"A{SC}:L{SC}")
        ws[f"A{SC}"].value = f"▌ 波段交易真實回測 (True Walk-Forward / 包含 {COST_PCT}% 交易成本)"
        ws[f"A{SC}"].font  = Font(bold=True, size=11, color="FFFFFF", name="Arial")
        ws[f"A{SC}"].fill  = PatternFill("solid", fgColor="C0392B")

        for ci, h in enumerate(["回測區間", "決策截斷日", "預測漲幅(%)", "單純原始漲幅", "扣除成本實戰績效", "實戰方向判定"], 1):
            cell = ws.cell(row=SC + 1, column=ci, value=h)
            cell.style = "header_style"
            cell.fill = PatternFill("solid", fgColor="E74C3C")

        for ri, r in enumerate(val_records):
            for ci, v in enumerate([r["period"], r["entry_date"], r["pred_return"], f"{r['actual_return']}%", f"{r['strat_return']}%", r["is_hit"]], 1):
                cell = ws.cell(row=SC + 2 + ri, column=ci, value=v)
                cell.style = "basic_border"
                cell.fill = PatternFill("solid", fgColor="FDEDEC" if ri % 2 == 0 else "FADBD8")
                cell.font = Font(name="Arial", size=10, bold=(ci == 6), color="27AE60" if "✅" in str(v) else ("C0392B" if "❌" in str(v) else "000000"))

        SR = SC + len(val_records) + 3
        ws.merge_cells("A{}:D{}".format(SR, SR))
        ws[f"A{SR}"].value = "波段回測績效 (Swing-Trend Backtest Stats)"
        ws[f"A{SR}"].font  = Font(bold=True, name="Arial", size=10, color="FFFFFF")
        ws[f"A{SR}"].fill  = PatternFill("solid", fgColor="922B21")
        ws[f"A{SR}"].alignment = Alignment(horizontal="center")
        
        for ci, (label, val) in enumerate([("實戰方向命中率", val_data["hit_rate"]), ("年化 實戰 Sharpe", val_data["sharpe"]), ("最大回撤 MDD", f"{val_data['mdd']}%"),
                                            ("Walk-Forward折數", val_data.get("n_folds", "N/A")), ("策略報酬標準差", f"{val_data.get('return_std', 0)}%"),
                                            ("💎Profit Factor", format_profit_factor(val_data.get("profit_factor"))),
                                            ("💎Expectancy(單筆期望值)", f"{val_data.get('expectancy', 0.0):+.2f}%")], 1):
            ws.cell(row=SR + 1, column=ci * 2 - 1, value=label).font = Font(bold=True, name="Arial", size=10)
            ws.cell(row=SR + 1, column=ci * 2, value=val).font = Font(name="Arial", size=10, color="1F497D")

    set_col_widths(ws, {"A": 10, "B": 22, "C": 15, "D": 15, "E": 14, "F": 16, "G": 15, "H": 14, "I": 12, "J": 12, "K": 12, "L": 12, "M": 22, "N": 14})
    ws.freeze_panes = "A15"

def write_validation_sheet(wb, val_data_all: list):
    ws = wb.create_sheet(title="真實回測總表", index=0)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:K1")
    ws["A1"].value = f"📈 股票多因子篩選 — 歷史真實回測總表（含 {COST_PCT}% 交易成本 | {VERSION}）"
    ws["A1"].font  = Font(bold=True, size=15, color="FFFFFF", name="Arial")
    ws["A1"].fill  = PatternFill("solid", fgColor="8E44AD")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    for ci, h in enumerate(["股號", "公司名稱", "進場日", "出場日", "持股天數", "出場原因", "預測漲幅(%)", "單純發生漲幅", "扣除成本實戰績效", "實戰方向判定"], 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.style = "header_style"
        cell.fill = PatternFill("solid", fgColor="9B59B6")
    ws.row_dimensions[2].height = 22

    row_cursor = 3
    for d in val_data_all:
        for r in d["records"]:
            for ci, v in enumerate([d["stock_id"], d["company"], r["entry_date"], r["exit_date"], f"{r['hold_days']}天", r["reason"], r["pred_return"], f"{r['actual_return']}%", f"{r['strat_return']}%", r["is_hit"]], 1):
                cell = ws.cell(row=row_cursor, column=ci, value=v)
                cell.style = "basic_border"
                cell.fill = PatternFill("solid", fgColor="F5EEF8" if row_cursor % 2 == 0 else "EBDEF0")
                cell.font = Font(name="Arial", size=10, bold=(ci == 10), color="27AE60" if "✅" in str(v) else ("C0392B" if "❌" in str(v) else "000000"))
            row_cursor += 1

    all_records = [r for d in val_data_all for r in d["records"]]
    if all_records:
        total_hits = sum(1 for r in all_records if r["is_hit_bool"])
        overall_rate = f"{total_hits}/{len(all_records)} ({total_hits/len(all_records)*100:.0f}%)"
        agg_stats = compute_trade_stats(all_records)
        ws.cell(row=row_cursor + 2, column=1, value="整體實戰命中率").font = Font(bold=True, name="Arial")
        ws.cell(row=row_cursor + 2, column=2, value=overall_rate).font = Font(name="Arial", color="1F497D", bold=True)
        ws.cell(row=row_cursor + 3, column=1, value="整體 Profit Factor").font = Font(bold=True, name="Arial")
        ws.cell(row=row_cursor + 3, column=2, value=format_profit_factor(agg_stats.get("profit_factor"))).font = Font(name="Arial", color="27AE60", bold=True)
        ws.cell(row=row_cursor + 4, column=1, value="整體 Expectancy(單筆期望值)").font = Font(bold=True, name="Arial")
        ws.cell(row=row_cursor + 4, column=2, value=f"{agg_stats.get('expectancy', 0.0):+.2f}%").font = Font(name="Arial", color="27AE60", bold=True)
        ws.cell(row=row_cursor + 5, column=1, value="Gross Profit / Gross Loss").font = Font(bold=True, name="Arial")
        ws.cell(row=row_cursor + 5, column=2, value=f"{agg_stats.get('gross_profit',0):.1f}% / {agg_stats.get('gross_loss',0):.1f}%").font = Font(name="Arial", color="1F497D")

    set_col_widths(ws, {"A": 10, "B": 15, "C": 12, "D": 16, "E": 16, "F": 16, "G": 18, "H": 14, "I": 16, "J": 12})
    ws.freeze_panes = "A3"

def write_summary_sheet(wb, summary_data: list, val_data_all: list, tab_name: str):
    ws = wb.create_sheet(title="投資建議總表", index=1)
    ws.sheet_view.showGridLines = False
    
    # 復原 image_d80821.jpg 列 1 與列 2 格式
    ws.merge_cells("A1:W1")
    ws["A1"].value = f"🏆 法人級多週期綜合排名推薦（{VERSION} 實戰資金控管波段版）"
    ws["A1"].font  = Font(bold=True, size=15, color="FFFFFF", name="Arial")
    ws["A1"].fill  = PatternFill("solid", fgColor=COLORS["header_bg"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # === 🚀 核心升級：增設波段量化總體表現看板 (Executive Portfolio Dashboard) ===
    all_records = [r for d in val_data_all for r in d["records"]]
    if all_records:
        total_trades = len(all_records)
        total_hits = sum(1 for r in all_records if r["is_hit_bool"])
        win_rate = total_hits / total_trades * 100
        avg_hold = np.mean([r["hold_days"] for r in all_records])
        avg_ret = np.mean([r["strat_return"] for r in all_records])
        agg_stats = compute_trade_stats(all_records)   # V18.4：整體 Profit Factor / Expectancy
    else:
        total_trades, win_rate, avg_hold, avg_ret = 0, 0.0, 0.0, 0.0
        agg_stats = compute_trade_stats([])

    dash_style = Font(bold=True, size=11, color="1F3864", name="Arial")
    val_style = Font(bold=True, size=16, color="C0392B" if avg_ret < 0 else "27AE60", name="Arial")
    center_align = Alignment(horizontal="center", vertical="center")
    bg_fill = PatternFill("solid", fgColor="F2F4F4")

    # B3:C4 -> 整體波段勝率
    ws.merge_cells("B3:C3")
    ws["B3"] = "🎯 整體波段實戰勝率"
    ws["B3"].font, ws["B3"].alignment, ws["B3"].fill = dash_style, center_align, bg_fill
    ws.merge_cells("B4:C4")
    ws["B4"] = f"{win_rate:.1f}%"
    ws["B4"].font, ws["B4"].alignment, ws["B4"].fill = Font(bold=True, size=16, color="1F3864", name="Arial"), center_align, bg_fill

    # E3:F3 -> 平均抱股天數 (讓獲利奔跑)
    ws.merge_cells("E3:F3")
    ws["E3"] = "⏳ 平均波段抱股天數"
    ws["E3"].font, ws["E3"].alignment, ws["E3"].fill = dash_style, center_align, bg_fill
    ws.merge_cells("E4:F4")
    ws["E4"] = f"{avg_hold:.1f} 天"
    ws["E4"].font, ws["E4"].alignment, ws["E4"].fill = Font(bold=True, size=16, color="2E75B6", name="Arial"), center_align, bg_fill

    # H3:I3 -> 平均單筆波段交易報酬（=Expectancy 每筆交易平均期望值，專家建議比勝率更重要的核心指標）
    ws.merge_cells("H3:I3")
    ws["H3"] = "📈 Expectancy 每筆期望值"
    ws["H3"].font, ws["H3"].alignment, ws["H3"].fill = dash_style, center_align, bg_fill
    ws.merge_cells("H4:I4")
    ws["H4"] = f"{avg_ret:+.2f}%"
    ws["H4"].font, ws["H4"].alignment, ws["H4"].fill = val_style, center_align, bg_fill

    # K3:L3 -> 波段交易總次數
    ws.merge_cells("K3:L3")
    ws["K3"] = "📊 歷史波段交易總次數"
    ws["K3"].font, ws["K3"].alignment, ws["K3"].fill = dash_style, center_align, bg_fill
    ws.merge_cells("K4:L4")
    ws["K4"] = f"{total_trades} 次"
    ws["K4"].font, ws["K4"].alignment, ws["K4"].fill = Font(bold=True, size=16, color="7D3C98", name="Arial"), center_align, bg_fill

    # N3:O3 -> 【V18.4新增】Profit Factor（Gross Profit / Gross Loss，專家建議比勝率更重要）
    pf_val = agg_stats.get("profit_factor")
    pf_display = format_profit_factor(pf_val)
    pf_color = "27AE60" if (pf_val is not None and pf_val >= 1.5) else ("C0392B" if (pf_val is not None and pf_val < 1.0) else "F39C12")
    ws.merge_cells("N3:O3")
    ws["N3"] = "💎 Profit Factor"
    ws["N3"].font, ws["N3"].alignment, ws["N3"].fill = dash_style, center_align, bg_fill
    ws.merge_cells("N4:O4")
    ws["N4"] = pf_display
    ws["N4"].font, ws["N4"].alignment, ws["N4"].fill = Font(bold=True, size=16, color=pf_color, name="Arial"), center_align, bg_fill

    # 繪製看板框線
    for r in range(3, 5):
        for c_idx in [2, 3, 5, 6, 8, 9, 11, 12, 14, 15]:
            ws.cell(row=r, column=c_idx).border = Border(left=Side(style="thin", color="BDD7EE"), right=Side(style="thin", color="BDD7EE"),
                                                         top=Side(style="thin", color="BDD7EE"), bottom=Side(style="thin", color="BDD7EE"))

    ws.merge_cells("A6:W6")
    ws["A6"].value = (f"分析日期：{datetime.now().strftime('%Y-%m-%d')}  |  綜合分數：5D/20D/勝率加權  |  排序基準：FinalScore(波段綜合分數×類股熱度)  |  "
                       f"Gross Profit:{agg_stats.get('gross_profit',0):.1f}% Gross Loss:{agg_stats.get('gross_loss',0):.1f}%（Profit Factor = Gross Profit ÷ Gross Loss，>=1.5通常視為穩健策略）")
    ws["A6"].font  = Font(italic=True, size=10, color="666666", name="Arial")
    ws["A6"].alignment = Alignment(horizontal="center")

    # 完全重新規劃為「波段決策」導向的欄位結構！徹底消除短線思維
    headers = ["綜合排名", "股號", "公司名稱", "最新日期", "目前收盤", 
               "資金配比", "硬停損防守價 (-6%)", "10EMA 防守價 (移動停利)", "近5日最低價 (最後支撐)", 
               "5日預測動能", "20日預測趨勢", "AI上漲機率", "歷史波段勝率", "平均抱股天數", 
               "市場環境", "趨勢評級", "交易信號", "波段綜合分數", "實戰Sharpe", 
               "🌐風險模式", "風險分數(0-100)", "VIX", "NASDAQ昨日", "SOX昨日",
               "第1名指標", "第2名指標",
               "🐎Horse星等", "HorseScore", "⭐AI信心等級", "💰Kelly建議倉位", "綜合建議倉位", "🎯Final Decision",
               "波段操作建議 (極重要)", "個股工作表",
               "💎Profit Factor", "💎Expectancy(單筆期望值)"]
    
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=7, column=ci, value=h)
        cell.style = "header_style"
    ws.row_dimensions[7].height = 22

    sig_bg = {"強力買入": "C8E6C9", "買入": "DCEDC8", "觀望": "EAEDED", "持平觀望": "FFF9C4", "減碼": "FFCCBC", "賣出": "FFCDD2", "🔥 黑馬起漲": "FFF2CC"}
    sig_fg = {"強力買入": "1B5E20", "買入": "33691E", "觀望": "34495E", "持平觀望": "F57F17", "減碼": "BF360C", "賣出": "B71C1C", "🔥 黑馬起漲": "D35400"}

    for ri, d in enumerate(summary_data):
        sig, up_p = d.get("signal", ""), d.get("up_prob", 50.0)
        bg, fg = sig_bg.get(sig, "FFFFFF"), sig_fg.get(sig, "000000")
        
        risk_regime_d = d.get("risk_regime", "NORMAL")
        regime_emoji_d = {"NORMAL": "🟢正常", "CAUTION": "🟡警戒", "HIGH_RISK": "🔴高風險", "CRASH": "⚫崩盤"}.get(risk_regime_d, "⚪未知")
        vals = [
            d.get("rank", ri+1), 
            d.get("stock_id", ""), 
            d.get("company", ""), 
            d.get("latest_date", ""), 
            d.get("current_price", ""),
            d.get("position_size", ""), 
            d.get("stop_loss", ""), 
            d.get("trailing_stop", ""), # 10EMA 價位
            d.get("low_5", ""),         # 近5日低點
            f"{d.get('pred_return', 0):+.2f}%", 
            f"{d.get('pred_20d', 0):+.2f}%",
            f"{up_p:.1f}%", 
            d.get("val_hit_rate", "N/A"), 
            f"{d.get('val_avg_hold', 0.0):.1f} 天",
            d.get("regime_raw", "N/A"), 
            d.get("trend_stars", ""), 
            sig, 
            d.get("composite_score", ""),
            d.get("sharpe", 0.0),
            regime_emoji_d,
            d.get("risk_score", 0.0),
            d.get("vix_close", 18.0),
            f"{d.get('nasdaq_ret1', 0.0):+.2f}%",
            f"{d.get('sox_ret1', 0.0):+.2f}%",
            d.get("top1_feature", ""), 
            d.get("top2_feature", ""), 
            d.get("horse_stars", "★☆☆☆☆"),   # V18.1：改為星等分級，取代 YES/NO
            d.get("horse_score", 0),
            d.get("confidence_stars", ""),
            d.get("kelly_str", "N/A"),
            d.get("final_position_blend", ""),
            d.get("final_decision", ""),
            ("🚧" + d.get("advice", "")) if d.get("gate_triggered") else d.get("advice", ""), 
            d.get("sheet_name", ""),
            format_profit_factor(d.get("profit_factor")),
            f"{d.get('expectancy', 0.0):+.2f}%"
        ]
        
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=8 + ri, column=ci, value=v)
            cell.style = "basic_border"
            cell.fill = PatternFill("solid", fgColor=(("FFF2CC" if d.get("horse_flag") else bg) if ci == 27 else bg))
            
            cell.font = Font(name="Arial", size=10, bold=(ci in [1, 2, 3, 6, 7, 8, 9, 17, 18, 27, 29, 30]), color=fg if ci == 17 else ("2874A6" if ci == 6 else ("D35400" if ci == 27 and d.get("horse_flag") else "000000")))
            
            if ci in [18]: cell.number_format = "0.00"

    if len(summary_data) > 0:
        ws.conditional_formatting.add(f"R8:R{7 + len(summary_data)}", ColorScaleRule(start_type="min", start_color="FFCDD2", mid_type="num", mid_value=0, mid_color="FFFFFF", end_type="max", end_color="C8E6C9"))
    
    set_col_widths(ws, {
        "A": 10, "B": 10, "C": 14, "D": 12, "E": 12, "F": 12, "G": 18, "H": 22, "I": 22, "J": 14, "K": 14, "L": 14,
        "M": 16, "N": 16, "O": 16, "P": 14, "Q": 14, "R": 14, "S": 14, "T": 18, "U": 18, "V": 45, "W": 15,
        # V18 / V18.1 新增欄位（AA~AH）
        "AA": 14, "AB": 12, "AC": 14, "AD": 16, "AE": 16, "AF": 30, "AG": 45, "AH": 15,
        # V18.4 新增欄位（Profit Factor / Expectancy）
        "AI": 16, "AJ": 20
    })
    ws.freeze_panes = "A8"

# ==========================================
# 🐎 V18：Black Horse Rank 黑馬排行榜（獨立分頁，只列最有爆發力的股票）
# ==========================================
def write_horse_rank_sheet(wb, summary_data: list):
    horse_candidates = [d for d in summary_data if d.get("horse_flag")]
    horse_candidates.sort(key=lambda x: x.get("horse_score", 0), reverse=True)
    horse_candidates = horse_candidates[:HORSE_RANK_TOP_N]

    ws = wb.create_sheet(title="黑馬排行榜", index=2)
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:M1")
    ws["A1"].value = f"🔥 Black Horse Rank｜黑馬候選股排行榜（HorseScore >= {HORSE_SCORE_THRESHOLD} 且僅列前 {HORSE_RANK_TOP_N} 名｜{VERSION}）"
    ws["A1"].font  = Font(bold=True, size=15, color="FFFFFF", name="Arial")
    ws["A1"].fill  = PatternFill("solid", fgColor="D35400")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:M2")
    ws["A2"].value = ("說明：HorseScore 為獨立於 AI 模型分數之外的黑馬評分機制（BB擠壓+20／通道窄<10%分位+15／"
                       "股數爆量+15／成交值爆量>2.5倍(資金流)+15／20日新高+15／MACD金叉+10／ADX>25+10／法人連買+10），"
                       "僅供辨識短期爆發力，不直接影響 AI 上漲機率與波段綜合分數。"
                       "Horse星等：90+★★★★★爆發黑馬／80+★★★★☆強黑馬／70+★★★☆☆准黑馬／50+★★☆☆☆中等／<50★☆☆☆☆普通。")
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws["A2"].alignment = Alignment(horizontal="left", wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 28

    headers = ["Horse排名", "股號", "公司名稱", "HorseScore", "Horse星等", "交易信號", "AI信心等級",
               "波段綜合分數", "5日預測動能", "命中訊號組合", "市場環境", "🌐風險模式", "個股工作表"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.style = "header_style"
        cell.fill = PatternFill("solid", fgColor="E67E22")
    ws.row_dimensions[4].height = 22

    if not horse_candidates:
        ws.merge_cells("A5:M5")
        ws["A5"].value = "本次掃描未發現符合條件的黑馬候選股（HorseScore 未達門檻）。"
        ws["A5"].font = Font(italic=True, size=11, color="999999", name="Arial")
        ws["A5"].alignment = Alignment(horizontal="center", vertical="center")
    else:
        for ri, d in enumerate(horse_candidates, 1):
            risk_regime_d = d.get("risk_regime", "NORMAL")
            regime_emoji_d = {"NORMAL": "🟢正常", "CAUTION": "🟡警戒", "HIGH_RISK": "🔴高風險", "CRASH": "⚫崩盤"}.get(risk_regime_d, "⚪未知")
            vals = [
                ri, d.get("stock_id", ""), d.get("company", ""), d.get("horse_score", 0), d.get("horse_stars", "★☆☆☆☆"),
                d.get("signal", ""), d.get("confidence_stars", ""), d.get("composite_score", ""),
                f"{d.get('pred_return', 0):+.2f}%", d.get("horse_breakdown_str", ""),
                d.get("regime_raw", "N/A"), regime_emoji_d, d.get("sheet_name", "")
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=4 + ri, column=ci, value=v)
                cell.style = "basic_border"
                cell.fill = PatternFill("solid", fgColor="FEF5E7" if ri % 2 == 0 else "FDEBD0")
                cell.font = Font(name="Arial", size=10, bold=(ci in [1, 4, 5]), color="D35400" if ci in [4, 5] else "000000")

        ws.conditional_formatting.add(
            f"D5:D{4 + len(horse_candidates)}",
            ColorScaleRule(start_type="min", start_color="FFF2CC", mid_type="percentile", mid_value=50, mid_color="F5B041",
                            end_type="max", end_color="D35400")
        )

    set_col_widths(ws, {"A": 10, "B": 10, "C": 16, "D": 12, "E": 14, "F": 14, "G": 14, "H": 14, "I": 14, "J": 40, "K": 14, "L": 12, "M": 16})
    ws.freeze_panes = "A5"


def write_sector_heat_sheet(wb, sector_heat_map: dict, sector_heat_detail: dict, summary_data: list):
    """
    V18.3：資金輪動 Sector Heat 分頁 —— 列出本次監控清單中各細分類股的熱度排行，
    以及該類股中「個股 FinalScore」排名前幾名的成分股，方便直接看出目前資金往哪個主題輪動。
    """
    ranked = sorted(sector_heat_map.items(), key=lambda kv: kv[1], reverse=True)

    ws = wb.create_sheet(title="Sector_Heat資金輪動", index=3)
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    ws["A1"].value = f"🔥 資金輪動 Sector Heat Score（{VERSION}）"
    ws["A1"].font  = Font(bold=True, size=15, color="FFFFFF", name="Arial")
    ws["A1"].fill  = PatternFill("solid", fgColor="8E44AD")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    ws["A2"].value = ("說明：熱度分數(0~100，50為中性)以本次監控清單中，同細分類股近5日「價格動能＋量能擴張倍率」相對其他類股的"
                       "z-score正規化計算；>=70視為資金流入熱區🔥，<35視為資金退潮警示❄️。個股 FinalScore = 個股綜合分數 × 類股熱度倍率(0.6~1.4)。"
                       "分類樣本(N)過少(<2檔)時，熱度以中性50分處理，僅供參考。")
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws["A2"].alignment = Alignment(horizontal="left", wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 30

    headers = ["排名", "類股", "熱度分數", "樣本數N", "近5日均漲跌%", "均量擴張倍率", "熱區狀態", "該類股FinalScore領先個股"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.style = "header_style"
        cell.fill = PatternFill("solid", fgColor="8E44AD")
    ws.row_dimensions[4].height = 22

    if not ranked:
        ws.merge_cells("A5:H5")
        ws["A5"].value = "本次掃描樣本不足，無法計算類股熱度。"
        ws["A5"].font = Font(italic=True, size=11, color="999999", name="Arial")
        ws["A5"].alignment = Alignment(horizontal="center", vertical="center")
    else:
        for ri, (tag, score) in enumerate(ranked, 1):
            d = sector_heat_detail.get(tag, {})
            status = "🔥資金流入熱區" if score >= SECTOR_HEAT_HOT_THRESHOLD else ("❄️資金退潮警示" if score < SECTOR_HEAT_COLD_THRESHOLD else "➖中性")
            members = [m for m in summary_data if m.get("sector_tag") == tag]
            members.sort(key=lambda x: x.get("final_score_with_heat", 0), reverse=True)
            top_members_str = "、".join([f"{m.get('stock_id','')}{m.get('company','')}" for m in members[:3]]) or "N/A"
            vals = [ri, tag, score, d.get("count", 0), f"{d.get('avg_ret5', 0):+.2f}%",
                    f"{d.get('avg_vol_ratio', 1):.2f}x", status, top_members_str]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=4 + ri, column=ci, value=v)
                cell.style = "basic_border"
                cell.fill = PatternFill("solid", fgColor="F4ECF7" if ri % 2 == 0 else "E8DAEF")
                cell.font = Font(name="Arial", size=10, bold=(ci in [1, 2, 3]))

        ws.conditional_formatting.add(
            f"C5:C{4 + len(ranked)}",
            ColorScaleRule(start_type="min", start_color="AED6F1", mid_type="num", mid_value=50, mid_color="F5EEF8",
                            end_type="max", end_color="D35400")
        )

    set_col_widths(ws, {"A": 8, "B": 12, "C": 12, "D": 10, "E": 14, "F": 14, "G": 16, "H": 40})
    ws.freeze_panes = "A5"


# ==========================================
# 🚀 輔助函式：自動建立 Excel 輸出檔名
# ==========================================
def build_output_filename(tab_name: str) -> str:
    safe_tab = tab_name
    for ch in INVALID_SHEET_CHARS:
        safe_tab = safe_tab.replace(ch, "")
    today = datetime.now().strftime("%Y%m%d")
    base  = f"股票量化報告_{safe_tab}_{VERSION}_{today}"
    seq = 1
    while os.path.exists(f"{base}_{seq:02d}.xlsx"): 
        seq += 1
    return f"{base}_{seq:02d}.xlsx"

def run_analysis(gid: str = None, tab_name: str = None, custom_codes: str = None,
                  sheet_id: str = None, output_dir: str = ".", progress_cb=None):
    """
    可重複呼叫的分析入口（供 Gradio 等外部介面使用）。

    參數：
      gid          Google Sheet 分頁 gid（與 custom_codes 擇一使用；custom_codes 優先）
      tab_name     報表/分頁顯示用的名稱（例如「權值股」「自選股」）
      custom_codes 自選股股號字串（例如 "2330,2317,0050"），有值時優先於 gid 讀 Google Sheet
      sheet_id     覆寫 Google Sheet 檔案 ID（預設用模組內建 SHEET_ID）
      output_dir   Excel 報告輸出資料夾
      progress_cb  進度回呼 function(idx:int, total:int, stock_id:str, company:str, message:str)

    回傳：(output_file_path, summary_data, val_data_all) — 任一步驟失敗則回傳 (None, [], [])
    """
    tab_name = tab_name or SHEET_TAB_NAME

    def _report(idx, total, sid, comp, msg):
        print(msg)
        if progress_cb:
            try:
                progress_cb(idx, total, sid, comp, msg)
            except Exception:
                pass

    print("="*65)
    print(f"🚀 啟動 整合型量化分析引擎 ({VERSION} | Horse星等 + AI Gate動態門檻 + MARKET_MODE + Sector Heat資金輪動 + Kelly Fallback + Final Decision Engine)")
    print(f"    分析清單：{tab_name}")
    print("="*65)

    if custom_codes and str(custom_codes).strip():
        stocks = build_custom_stock_list(custom_codes)
    else:
        stocks = load_stock_list(gid=gid, sheet_id=sheet_id)
    if not stocks:
        print("❌ 股票清單為空，中止分析。")
        return None, [], []

    wb = openpyxl.Workbook()
    for style_name in ["header_style", "basic_border"]:
        if style_name in wb.named_styles: del wb.named_styles[style_name]
    setup_excel_styles(wb)
    if "Sheet" in wb.sheetnames: del wb["Sheet"]

    try:
        dl = DataLoader()
        if FINMIND_TOKEN: dl.login_by_token(FINMIND_TOKEN)
    except Exception as e:
        print(f"⚠️ FinMind 初始化失敗: {e}")
        return None, [], []

    print("\n📦 載入大盤環境變數...")
    df_macro = process_macro_data(fetch_yahoo_data("^TWII"))
    if not df_macro.empty: print(f"  ✅ 成功載入大盤環境變數 ({len(df_macro)} 筆)")

    # ========= V15：下載全球風險因子 =========
    print("\n🌐 V15 全球風險覆寫層：下載美股風險指標...")
    us_risk = fetch_us_risk_data()
    global_risk_score, global_risk_regime, global_risk_detail = compute_global_risk_score(us_risk)
    regime_emoji = {"NORMAL": "🟢", "CAUTION": "🟡", "HIGH_RISK": "🔴", "CRASH": "⚫"}.get(global_risk_regime, "⚪")
    pos_limit_map = {"NORMAL": POSITION_SIZE_BASE, "CAUTION": POSITION_SIZE_CAUTION,
                     "HIGH_RISK": POSITION_SIZE_HIGHRISK, "CRASH": POSITION_SIZE_CRASH}
    pos_limit_now = pos_limit_map.get(global_risk_regime, POSITION_SIZE_BASE)
    print(f"  {regime_emoji} 全球風險模式：{global_risk_regime} | 量化分數：{global_risk_score:.1f}/100 | 倉位上限：{pos_limit_now*100:.0f}%")
    print(f"  📊 風險細節：{global_risk_detail}")

    # ========= V18.2b：市場 Regime 分類 MARKET_MODE（全站banner，個股仍各自依 macro_regime 對齊計算） =========
    _macro_regime_now = df_macro["macro_regime"].iloc[-1] if (not df_macro.empty and "macro_regime" in df_macro.columns) else "低波動盤整"
    global_market_mode = classify_market_mode(_macro_regime_now, us_risk.get("vix_close") if us_risk.get("fetch_ok") else None)
    _mode_pb = MARKET_MODE_PLAYBOOK.get(global_market_mode, {})
    print(f"  {_mode_pb.get('label','')} MARKET_MODE：{global_market_mode}（大盤regime={_macro_regime_now}）"
          f" | 偏好風格：{_mode_pb.get('favored_style','')} | 偏好類股：{'/'.join(_mode_pb.get('favored_categories', [])) or '無(全面防禦)'}")

    # ========= V18.3：資金輪動 Sector Heat Score（輕量預掃，同時建立 price_cache 供主迴圈重複使用） =========
    print("\n🔥 V18.3 資金輪動層：計算類股 Sector Heat Score（AI/半導體/PCB/散熱/網通/航運/金融/傳產...）...")
    price_cache = {}
    sector_heat_map, sector_heat_detail = compute_sector_heat_scores(stocks, price_cache)
    if sector_heat_map:
        ranked_sectors = sorted(sector_heat_map.items(), key=lambda kv: kv[1], reverse=True)
        heat_lines = []
        for tag, score in ranked_sectors:
            d = sector_heat_detail.get(tag, {})
            tag_emoji = "🔥" if score >= SECTOR_HEAT_HOT_THRESHOLD else ("❄️" if score < SECTOR_HEAT_COLD_THRESHOLD else "➖")
            heat_lines.append(f"{tag_emoji}{tag}:{score:.0f}分(N={d.get('count',0)},5日均漲{d.get('avg_ret5',0):+.1f}%)")
        print("  " + " | ".join(heat_lines))
    else:
        print("  ⚠️ 樣本不足或資料下載失敗，Sector Heat Score 暫以中性 50 分處理。")

    summary_data, val_data_all = [], []
    total_stocks = len(stocks)

    for idx, (ticker, stock_id, company) in enumerate(stocks, 1):
        _report(idx, total_stocks, stock_id, company, f"\n[{idx}/{total_stocks}] 正在分析 {stock_id} {company} ...")
        
        df_price = price_cache.get(stock_id)
        if df_price is None:
            df_price = fetch_yahoo_data(ticker)
            price_cache[stock_id] = df_price
        if df_price is None or df_price.empty:
            print(f"  ⚠️ 無法取得價量資料，跳過。")
            continue
            
        df_chips = fetch_chip_data(stock_id, dl)
        df = df_price.copy()
        if not df_chips.empty:
            df = df.join(df_chips, how='left')
            for col in ['Foreign_Net', 'Trust_Net']:
                if col in df.columns: df[col] = df[col].fillna(0)
            for col in ['Margin_Bal', 'Short_Bal']:
                if col in df.columns: df[col] = df[col].ffill().fillna(0)
                
        if len(df) < 80:
            print(f"  ⚠️ 資料筆數不足 ({len(df)} < 80 筆)，略過。")
            continue

        if df['volume'].tail(20).mean() < (1000 * 1000 if df['volume'].mean() > 50000 else 1000):
            print(f"  ⚠️ 流動性不足，略過。")
            continue

        if not df_macro.empty:
            df = df.join(df_macro, how='left')
            for col in ["macro_sma20", "macro_sma60", "macro_bias20", "macro_macd_hist", "macro_close", "macro_ret20", "macro_ret1d", "macro_regime", "macro_regime_score", "macro_ADL_trend"]:
                if col in df.columns: df[col] = df[col].ffill()

        df = compute_indicators(df)
        df = define_target(df)
        df.replace([np.inf, -np.inf], np.nan, inplace=True)

        all_features, effective = feature_selection(df)
        if all_features is None: 
            print("  ⚠️ 特徵或資料筆數不足，略過預測")
            continue

        top_features_list = effective["feature"].tolist()
        df_tune = df[top_features_list + ["target_5d"]].dropna()
        best_params = optimize_hyperparameters(df_tune[top_features_list], df_tune["target_5d"]) if len(df_tune) > 100 else {'max_depth': 4, 'learning_rate': 0.05, 'subsample': 0.8}

        # ================= V18.2：類股分類（供動態 AI Gate 門檻 / 分業風險模型使用） =================
        stock_category = classify_stock_category(stock_id, company)

        pred = predict_with_top_features(df, top_features_list, best_params, us_risk=us_risk, stock_category=stock_category)
        print(f"  💡 [{stock_category}] {pred['advice']}")

        val_data = historical_validation(df, top_features_list, best_params)
        
        # 💡 計算該股在 Walk-Forward 中真實的動態波段勝率與抱股時間
        val_records = val_data.get("records", [])
        if val_records:
            stock_val_hits = sum(1 for r in val_records if r["is_hit_bool"])
            stock_val_winrate = f"{stock_val_hits}/{len(val_records)} ({stock_val_hits/len(val_records)*100:.0f}%)"
            stock_val_avg_hold = np.mean([r["hold_days"] for r in val_records])
        else:
            stock_val_winrate = "N/A"
            stock_val_avg_hold = 0.0

        # ================= V18：HorseFinder 黑馬評分（獨立機制，不影響 AI 分數） =================
        horse_result = compute_horse_score(df)
        if horse_result["horse_flag"]:
            print(f"  🔥 黑馬候選股！HorseScore={horse_result['horse_score']} ｜ {horse_breakdown_to_str(horse_result['breakdown'])}")

        # ================= V18.2：AI Decision Gate（動態門檻：類股別 + 趨勢分數 + 歷史Sharpe + 市場模式） =================
        market_mode = pred.get("market_mode", "SIDEWAYS")   # V18.2b
        dynamic_gate_threshold = compute_dynamic_ai_gate_threshold(
            stock_category, pred.get("trend_score", 0.0), val_data.get("sharpe", 0.0), market_mode=market_mode
        )
        gate_result = apply_ai_decision_gate(
            pred.get("signal"), pred.get("advice"), pred.get("prob_5d", 50.0), horse_result,
            ai_gate_threshold=dynamic_gate_threshold, category=stock_category
        )
        pred["original_signal"] = pred.get("signal")   # 保留 Gate 前的原始訊號，供報表對照
        pred["original_advice"] = pred.get("advice")
        pred["signal"] = gate_result["signal"]
        pred["advice"] = gate_result["advice"]
        pred["gate_triggered"] = gate_result["gate_triggered"]
        pred["gate_threshold"] = gate_result.get("gate_threshold", dynamic_gate_threshold)
        if gate_result["gate_triggered"]:
            print(f"  🚧 AI Decision Gate 觸發(門檻{dynamic_gate_threshold:.0f}%)：{pred['original_signal']} → {pred['signal']}")

        # ================= V18.2b：市場 Regime 分類 MARKET_MODE 風格過濾器 =================
        # 不同市場適合不同策略：多頭→突破動能／盤整→均值回歸RSI低接／空頭→防守金融逆勢／恐慌→現金為王。
        # 避免「模型在空頭市場還硬做非金融股突破」這種風格與市場不match的情況。
        mode_result = apply_market_mode_filter(
            pred.get("signal"), pred.get("advice"), market_mode, stock_category, pred.get("latest_rsi", 50.0)
        )
        if mode_result["mode_adjusted"]:
            pred["advice"] = mode_result["advice"]
            if mode_result["signal"] != pred.get("signal"):
                print(f"  🎯 MARKET_MODE={market_mode} 風格過濾：{pred['signal']} → {mode_result['signal']}")
            pred["signal"] = mode_result["signal"]
        pred["mode_adjusted"] = mode_result["mode_adjusted"]

        # ================= V18.3：資金輪動 Sector Heat Score → FinalScore = 個股分數 × 類股熱度 =================
        sector_tag = classify_sector_fine(stock_id, company)
        sector_heat = sector_heat_map.get(sector_tag, 50.0)
        heat_multiplier = SECTOR_HEAT_MULTIPLIER_MIN + (sector_heat / 100.0) * (SECTOR_HEAT_MULTIPLIER_MAX - SECTOR_HEAT_MULTIPLIER_MIN)
        final_score_with_heat = round(pred.get("composite_score", 0.0) * heat_multiplier, 2)
        pred["sector_tag"] = sector_tag
        pred["sector_heat"] = sector_heat
        pred["final_score_with_heat"] = final_score_with_heat

        if sector_heat < SECTOR_HEAT_COLD_THRESHOLD and pred.get("signal") in ["強力買入", "買入", "🔥 黑馬起漲", "🚀 搶先布局(Horse Override)"]:
            pred["advice"] += (f" ｜⚠️ 資金輪動警示：所屬類股「{sector_tag}」熱度僅{sector_heat:.0f}分(資金退潮區，"
                                f"閾值<{SECTOR_HEAT_COLD_THRESHOLD})，資金可能已輪動至其他類股，追高風險較高，建議留意量能是否同步萎縮。")
        elif sector_heat >= SECTOR_HEAT_HOT_THRESHOLD:
            pred["advice"] += f" ｜🔥 資金輪動：所屬類股「{sector_tag}」熱度{sector_heat:.0f}分，屬本次資金流入熱區。"

        # ================= V18：Kelly Position Sizing（與 ATR / 風險倉位聯合取最小值） =================
        kelly_info = compute_kelly_position(val_records, prob_5d=pred.get("prob_5d", 50.0))
        atr_pos_raw    = pred.get("atr_position_raw", 0.20)
        regime_pos_raw = pred.get("regime_position_raw", atr_pos_raw)
        pos_candidates = [atr_pos_raw, regime_pos_raw]
        if kelly_info.get("kelly_fraction") is not None:
            pos_candidates.append(kelly_info["kelly_fraction"])
        final_position_raw = min(pos_candidates)
        final_position_str = f"{final_position_raw * 100:.1f}%" if final_position_raw > 0 else "⚫ 全現金"

        # ================= V18.1：AI 信心等級（★ 星等） =================
        confidence = compute_confidence_stars(pred, val_data, horse_result, df)

        # ================= V18.1：Final Decision Engine（統整 AI/Horse/Risk/Gate，避免報表欄位互相矛盾） =================
        final_decision = compute_final_decision(pred, horse_result, pred.get("risk_regime", "NORMAL"))

        sheet_name = clean_sheet_name(stock_id, company)
        write_stock_sheet(wb, sheet_name, stock_id, company, all_features, effective, df, pred, val_data,
                           horse_result=horse_result, kelly_info=kelly_info,
                           final_position_str=final_position_str, confidence=confidence, final_decision=final_decision)

        if val_data["records"]:
            val_data_all.append({"stock_id": stock_id, "company": company, "records": val_data["records"], "sharpe": val_data["sharpe"], "mdd": val_data["mdd"],
                                  "profit_factor": val_data.get("profit_factor"), "expectancy": val_data.get("expectancy", 0.0)})

        top3_eff = effective.head(3)
        summary_data.append({
            "stock_id": stock_id, "company": company, "sheet_name": sheet_name,
            "latest_date": pred.get("latest_date"), "current_price": pred.get("current_price"),
            "stop_loss": pred.get("stop_loss"), 
            "trailing_stop": pred.get("trailing_stop"), # 傳入 10EMA 移動防守價
            "low_5": round(float(df['Low_5'].iloc[-1]), 2) if 'Low_5' in df.columns else pred.get("current_price"), # 傳入近5日最低點
            "pred_1d": pred.get("pred_1d"), 
            "pred_3d": pred.get("pred_3d"),
            "pred_return": pred.get("pred_return"), 
            "pred_20d": pred.get("pred_20d"),
            "up_prob": pred.get("prob_5d"), 
            "position_size": pred.get("position_size"),
            "tp_cons": pred.get("tp_cons"), 
            "tp_agg": pred.get("tp_agg"),
            "val_hit_rate": stock_val_winrate, # 傳入個股回測勝率
            "val_avg_hold": stock_val_avg_hold, # 傳入個股平均抱股天數
            "regime_raw": df["macro_regime"].iloc[-1] if "macro_regime" in df.columns else "N/A",
            "trend_stars": pred.get("trend_stars"),
            "composite_score": pred.get("composite_score"), "sharpe": val_data.get("sharpe", 0.0),
            "signal": pred.get("signal"), "advice": pred.get("advice"),
            "original_signal": pred.get("original_signal"), "gate_triggered": pred.get("gate_triggered", False),
            "top1_feature": top3_eff.iloc[0]["feature"] if len(top3_eff) > 0 else "",
            "top2_feature": top3_eff.iloc[1]["feature"] if len(top3_eff) > 1 else "",
            "risk_score": pred.get("risk_score", 0),
            "risk_regime": pred.get("risk_regime", "NORMAL"),
            "risk_detail": pred.get("risk_detail", ""),
            "vix_close": pred.get("vix_close", 18.0),
            "nasdaq_ret1": pred.get("nasdaq_ret1", 0.0),
            "sox_ret1": pred.get("sox_ret1", 0.0),
            # ========= V18 新增欄位 =========
            "horse_score": horse_result["horse_score"],
            "horse_flag": horse_result["horse_flag"],
            "horse_stars": horse_result.get("horse_stars", "★☆☆☆☆"),
            "horse_tier": horse_result.get("horse_tier", "極弱"),
            "horse_breakdown": horse_result["breakdown"],
            "horse_breakdown_str": horse_breakdown_to_str(horse_result["breakdown"]),
            "kelly_str": kelly_info.get("kelly_str", "N/A"),
            "kelly_is_fallback": kelly_info.get("is_fallback", False),
            "final_position_blend": final_position_str,
            "final_decision": final_decision,
            "confidence_stars": confidence["stars"],
            "confidence_n": confidence["stars_n"],
            "bb_squeeze": int(df["BB_squeeze"].iloc[-1]) if "BB_squeeze" in df.columns else 0,
            # ========= V18.2 新增欄位 =========
            "stock_category": stock_category,
            "gate_threshold": pred.get("gate_threshold", AI_GATE_LOW_PROB),
            "trend_score": pred.get("trend_score", 0.0),
            "market_mode": pred.get("market_mode", "SIDEWAYS"),
            "mode_adjusted": pred.get("mode_adjusted", False),
            "sector_tag": pred.get("sector_tag", "其他"),
            "sector_heat": pred.get("sector_heat", 50.0),
            "final_score_with_heat": pred.get("final_score_with_heat", pred.get("composite_score", 0.0)),
            "profit_factor": val_data.get("profit_factor"), "expectancy": val_data.get("expectancy", 0.0),
            "gross_profit": val_data.get("gross_profit", 0.0), "gross_loss": val_data.get("gross_loss", 0.0),
            "turnover_ratio_20": round(float(df["Turnover_Ratio_20"].iloc[-1]), 2) if "Turnover_Ratio_20" in df.columns and pd.notna(df["Turnover_Ratio_20"].iloc[-1]) else 0.0,
            "macro_status": "GOOD" if (df["macro_regime"].iloc[-1] in ["強趨勢多頭", "緩漲盤"] if "macro_regime" in df.columns else False)
                            else ("BAD" if (df["macro_regime"].iloc[-1] in ["崩盤", "主跌段", "空頭反彈"] if "macro_regime" in df.columns else False) else "NEUTRAL"),
        })

    # V18.3：綜合排名改依 FinalScore(= composite_score × 類股熱度倍率) 排序，讓資金輪動熱區的股票排名更前面
    summary_data.sort(key=lambda x: x.get("final_score_with_heat") if x.get("final_score_with_heat") is not None else (x.get("composite_score") or -999), reverse=True)
    for rank, d in enumerate(summary_data, 1): d["rank"] = rank

    if val_data_all: write_validation_sheet(wb, val_data_all)
    write_summary_sheet(wb, summary_data, val_data_all, tab_name)
    write_horse_rank_sheet(wb, summary_data)   # V18：Black Horse Rank 黑馬排行榜
    write_sector_heat_sheet(wb, sector_heat_map, sector_heat_detail, summary_data)   # V18.3：資金輪動 Sector Heat

    os.makedirs(output_dir, exist_ok=True)
    safe_tab = tab_name
    for ch in INVALID_SHEET_CHARS:
        safe_tab = safe_tab.replace(ch, "")
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(output_dir, f"股票量化報告_{safe_tab}_{VERSION}_{stamp}.xlsx")
    wb.save(output_file)
    print(f"\n{'='*65}\n✅ 波段分析完成！\n資料皆在記憶體內完成運算，最終報告已儲存至：{output_file}\n{'='*65}")
    return output_file, summary_data, val_data_all


def main():
    """CLI 進入點：沿用模組頂部 SHEET_ID/GID/SHEET_TAB_NAME 設定執行一次分析。"""
    run_analysis(gid=GID, tab_name=SHEET_TAB_NAME, sheet_id=SHEET_ID)


if __name__ == "__main__":
    main()